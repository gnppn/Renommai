#!/usr/bin/env python3
"""
RenAIme - Script minimaliste de tri et renommage de documents
Basé sur OCR + Analyse IA Ollama + Llava Vision
"""

import os
import sys
import json
import csv
import re
import shutil
import tempfile
import subprocess
import signal
import atexit
import base64
from pathlib import Path
from datetime import datetime
from io import BytesIO

import pdfplumber
from PIL import Image, ImageEnhance, ImageFilter
import pytesseract
import ollama

try:
    from docx import Document
    DOCX_AVAILABLE = True
except:
    DOCX_AVAILABLE = False

try:
    from openpyxl import load_workbook
    XLSX_AVAILABLE = True
except:
    XLSX_AVAILABLE = False

try:
    from pypdf import PdfReader, PdfWriter
    PYPDF_AVAILABLE = True
except:
    PYPDF_AVAILABLE = False

try:
    import cv2
    import numpy as np
    CV2_AVAILABLE = True
except:
    CV2_AVAILABLE = False

try:
    from pdf2image import convert_from_path
    PDF2IMAGE_AVAILABLE = True
except:
    PDF2IMAGE_AVAILABLE = False

# ========== CONFIGURATION ==========

SCRIPT_DIR = Path(__file__).parent
PROMPTS_DIR = SCRIPT_DIR / "prompts"

# Cache pour la taille de contexte des modèles
_model_context_cache = {}

# Modèle Llava sélectionné (sera défini par detect_llava_model)
_selected_llava_model = None

def get_system_power_level():
    """Détecte la puissance du système et retourne un niveau (low, medium, high).
    
    Basé sur:
    - RAM disponible
    - Présence et VRAM GPU (si nvidia-smi disponible)
    """
    ram_gb = 0
    vram_gb = 0
    
    # Détecter la RAM système
    try:
        with open('/proc/meminfo', 'r') as f:
            for line in f:
                if line.startswith('MemTotal:'):
                    # MemTotal en kB
                    ram_kb = int(line.split()[1])
                    ram_gb = ram_kb / (1024 * 1024)
                    break
    except:
        ram_gb = 8  # Valeur par défaut
    
    # Détecter la VRAM GPU (NVIDIA)
    try:
        result = subprocess.run(
            ['nvidia-smi', '--query-gpu=memory.total', '--format=csv,noheader,nounits'],
            capture_output=True, text=True, timeout=5
        )
        if result.returncode == 0:
            # Prendre la première GPU, valeur en MB
            vram_mb = int(result.stdout.strip().split('\n')[0])
            vram_gb = vram_mb / 1024
    except:
        vram_gb = 0  # Pas de GPU NVIDIA ou nvidia-smi non disponible
    
    # Déterminer le niveau de puissance
    # High: >= 16GB RAM et >= 8GB VRAM, ou >= 32GB RAM
    # Medium: >= 8GB RAM et >= 4GB VRAM, ou >= 16GB RAM
    # Low: reste
    
    if (ram_gb >= 16 and vram_gb >= 8) or ram_gb >= 32 or vram_gb >= 12:
        return 'high', ram_gb, vram_gb
    elif (ram_gb >= 8 and vram_gb >= 4) or ram_gb >= 16 or vram_gb >= 6:
        return 'medium', ram_gb, vram_gb
    else:
        return 'low', ram_gb, vram_gb

def select_llava_model(available_models):
    """Sélectionne le meilleur modèle Llava selon la puissance du système.
    
    Versions Llava (du plus léger au plus lourd):
    - llava:7b - Le plus léger, pour machines limitées
    - llava:latest / llava:7b-v1.6 - Version standard
    - llava:13b - Plus performant, nécessite plus de ressources
    - llava:34b - Très performant, nécessite beaucoup de ressources
    """
    global _selected_llava_model
    
    power_level, ram_gb, vram_gb = get_system_power_level()
    
    # Filtrer les modèles llava disponibles
    llava_models = [m for m in available_models if m.startswith('llava')]
    
    # Définir les préférences selon le niveau de puissance
    if power_level == 'high':
        # Préférer les modèles plus gros si disponibles
        preferred = ['llava:34b', 'llava:13b', 'llava:13b-v1.6', 'llava:latest', 'llava:7b-v1.6', 'llava:7b']
        default_pull = 'llava:13b'
    elif power_level == 'medium':
        # Version standard
        preferred = ['llava:13b', 'llava:latest', 'llava:7b-v1.6', 'llava:7b']
        default_pull = 'llava:latest'
    else:
        # Version légère pour machines limitées
        preferred = ['llava:7b', 'llava:7b-v1.6', 'llava:latest']
        default_pull = 'llava:7b'
    
    # Chercher le meilleur modèle disponible
    for model in preferred:
        if model in llava_models:
            _selected_llava_model = model
            return model, power_level, ram_gb, vram_gb
    
    # Si aucun llava n'est disponible, on retourne celui à télécharger
    _selected_llava_model = default_pull
    return default_pull, power_level, ram_gb, vram_gb

def get_llava_model():
    """Retourne le modèle Llava sélectionné."""
    global _selected_llava_model
    return _selected_llava_model or 'llava:latest'

def get_model_context_size(model_name):
    """Récupère la taille de contexte d'un modèle Ollama."""
    if model_name in _model_context_cache:
        return _model_context_cache[model_name]
    
    try:
        result = subprocess.run(
            ['ollama', 'show', model_name],
            capture_output=True, text=True, timeout=10
        )
        output = result.stdout
        
        # Chercher "context length" ou "num_ctx" dans la sortie
        for line in output.splitlines():
            line_lower = line.lower()
            if 'context' in line_lower or 'num_ctx' in line_lower:
                # Extraire le nombre
                import re
                numbers = re.findall(r'\d+', line)
                if numbers:
                    ctx_size = int(numbers[-1])
                    _model_context_cache[model_name] = ctx_size
                    return ctx_size
        
        # Défaut si non trouvé: estimer selon le nom du modèle
        if '32k' in model_name or '32b' in model_name:
            return 32768
        elif '16k' in model_name:
            return 16384
        elif '8k' in model_name or '8b' in model_name:
            return 8192
        else:
            return 4096  # Défaut conservateur
    except:
        return 4096  # Défaut en cas d'erreur

def load_prompt(name, model_name=None):
    """Charge un prompt adapté à la taille du modèle."""
    if model_name:
        ctx_size = get_model_context_size(model_name)
        
        # Sélectionner le prompt selon la taille de contexte
        if ctx_size >= 16384:
            suffix = "_16k"
        elif ctx_size >= 8192:
            suffix = "_8k"
        else:
            suffix = "_4k"
        
        # Essayer le prompt spécifique, sinon le générique
        prompt_file = PROMPTS_DIR / f"{name}{suffix}.txt"
        if prompt_file.exists():
            return prompt_file.read_text(encoding='utf-8')
    
    # Fallback: prompt générique (4k par défaut)
    prompt_file = PROMPTS_DIR / f"{name}.txt"
    if prompt_file.exists():
        return prompt_file.read_text(encoding='utf-8')
    
    # Dernier recours: prompt 4k
    prompt_file = PROMPTS_DIR / f"{name}_4k.txt"
    if prompt_file.exists():
        return prompt_file.read_text(encoding='utf-8')
    
    raise FileNotFoundError(f"Prompt non trouvé: {name}")

DEFAULT_CONFIG = {
    "SOURCE_DIR": "documents",
    "EXPORT_DIR": "Export",
    "FAILURE_DIR": "Echec",
    "OLLAMA_MODEL": "llama3:8b-instruct-q4_0",
}

# Liste globale des fichiers temporaires à nettoyer
_temp_files = []

def cleanup_temp_files():
    """Nettoie tous les fichiers temporaires créés lors de l'exécution."""
    global _temp_files
    for tmp in _temp_files[:]:
        try:
            if os.path.exists(tmp):
                os.remove(tmp)
        except Exception as e:
            pass
        finally:
            _temp_files.remove(tmp) if tmp in _temp_files else None

def signal_handler(signum, frame):
    """Gestionnaire d'interruption (Ctrl+C, SIGTERM)."""
    signal_name = signal.Signals(signum).name
    print(f"\n\n⚠️  INTERRUPTION ({signal_name}) - Nettoyage en cours...")
    cleanup_temp_files()
    print("✅ Nettoyage terminé. Au revoir!")
    sys.exit(130)

# Enregistrer les gestionnaires de signaux
signal.signal(signal.SIGINT, signal_handler)   # Ctrl+C
signal.signal(signal.SIGTERM, signal_handler)  # Demande d'arrêt système

# Enregistrer la fonction de nettoyage pour appel automatique à la fermeture
atexit.register(cleanup_temp_files)

def load_config():
    """Charge ou crée la configuration."""
    if os.path.exists("config.json"):
        with open("config.json", "r", encoding="utf-8") as f:
            return {**DEFAULT_CONFIG, **json.load(f)}
    return DEFAULT_CONFIG

def save_config(config):
    """Sauvegarde la configuration."""
    with open("config.json", "w", encoding="utf-8") as f:
        json.dump(config, f, indent=2, ensure_ascii=False)

# ========== DÉPENDANCES ==========

def check_deps():
    """Vérifie les dépendances."""
    missing = []
    for pkg in ["pdfplumber", "PIL", "pytesseract", "ollama"]:
        try:
            __import__(pkg)
        except:
            missing.append(pkg)
    if missing:
        print(f"❌ Dépendances manquantes: {', '.join(missing)}")
        sys.exit(1)

def ensure_models():
    """Vérifie et télécharge les modèles Ollama manquants."""
    print("\n🤖 Vérification des modèles Ollama...")
    
    try:
        out = subprocess.run(['ollama', 'list'], stdout=subprocess.PIPE, text=True, check=True)
        available_models = [l.split()[0] for l in out.stdout.splitlines()[1:] if l.strip()]
        
        # Séparer modèles vision (llava) et texte (autres)
        text_models = [m for m in available_models if not m.startswith('llava')]
        llava_models = [m for m in available_models if m.startswith('llava')]
        
        if available_models:
            print("      ✅ Modèles disponibles")
        else:
            print("      ⚠️  Aucun modèle")
        
        # Sélectionner le modèle llava selon la puissance du PC
        selected_llava, power_level, ram_gb, vram_gb = select_llava_model(available_models)
        
        power_icons = {'high': '🚀', 'medium': '💻', 'low': '📱'}
        power_names = {'high': 'Élevée', 'medium': 'Moyenne', 'low': 'Limitée'}
        print(f"      {power_icons[power_level]} Puissance détectée: {power_names[power_level]} (RAM: {ram_gb:.1f}GB, VRAM: {vram_gb:.1f}GB)")
        print(f"      👁️  Modèle vision sélectionné: {selected_llava}")
        
        # Télécharger llava si absent
        if not llava_models or selected_llava not in llava_models:
            print(f"      ⬇️  Téléchargement {selected_llava}...")
            try:
                subprocess.run(['ollama', 'pull', selected_llava], check=True)
                print(f"      ✅ {selected_llava} téléchargé")
            except subprocess.CalledProcessError as e:
                print(f"      ❌ Échec: {e}")
                print(f"         Téléchargez manuellement: ollama pull {selected_llava}")
        
        # Télécharger llama3 seulement si AUCUN modèle texte n'est présent
        if not text_models:
            print("      ⬇️  Téléchargement llama3:8b-instruct-q4_0...")
            try:
                subprocess.run(['ollama', 'pull', 'llama3:8b-instruct-q4_0'], check=True)
                print("      ✅ llama3:8b-instruct-q4_0 téléchargé")
            except subprocess.CalledProcessError as e:
                print(f"      ❌ Échec: {e}")
                print("         Téléchargez manuellement: ollama pull llama3:8b-instruct-q4_0")
        
        print()
    
    except FileNotFoundError:
        print("      ❌ Ollama non installé")
        print("         https://ollama.ai")
        sys.exit(1)
    except Exception as e:
        print(f"      ⚠️  Erreur: {e}")

# ========== EXTRACTION TEXTE ==========

def preprocess_image_for_ocr(img):
    """Prétraitement image pour améliorer l'OCR."""
    img = img.convert('L')
    img = img.filter(ImageFilter.MedianFilter(size=3))
    
    if CV2_AVAILABLE:
        img_cv = cv2.cvtColor(np.array(img), cv2.COLOR_GRAY2BGR) if len(np.array(img).shape) == 2 else np.array(img)
        clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8, 8))
        img_cv_gray = cv2.cvtColor(img_cv, cv2.COLOR_BGR2GRAY)
        img_cv_gray = clahe.apply(img_cv_gray)
        img = Image.fromarray(img_cv_gray)
        
        img = ImageEnhance.Contrast(img).enhance(1.5)
        
        img_cv = cv2.cvtColor(np.array(img), cv2.COLOR_GRAY2BGR) if len(np.array(img).shape) == 2 else np.array(img)
        img_cv_gray = cv2.cvtColor(img_cv, cv2.COLOR_BGR2GRAY) if len(img_cv.shape) == 3 else img_cv
        _, img_cv_gray = cv2.threshold(img_cv_gray, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
        img = Image.fromarray(img_cv_gray)
        
        img_cv = np.array(img)
        kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (2, 2))
        img_cv = cv2.morphologyEx(img_cv, cv2.MORPH_CLOSE, kernel)
        img = Image.fromarray(img_cv)
    else:
        img = ImageEnhance.Contrast(img).enhance(1.5)
    
    img = img.filter(ImageFilter.SHARPEN)
    return img

def extract_from_pdf(path):
    """Extrait texte d'un PDF."""
    try:
        with pdfplumber.open(path) as pdf:
            texts = [page.extract_text() or "" for page in pdf.pages]
        return "\n".join(texts) if any(texts) else None
    except:
        return None

def create_searchable_pdf_page(img, vision_description=None):
    """Crée une page PDF avec image visible et couche OCR invisible."""
    try:
        pdf_img = BytesIO()
        img_rgb = img.convert('RGB')
        img_rgb.save(pdf_img, format='PDF')
        pdf_img.seek(0)
        
        pdf_ocr_bytes = pytesseract.image_to_pdf_or_hocr(img, extension='pdf')
        if not pdf_ocr_bytes:
            return pdf_img.getvalue()
        
        if PYPDF_AVAILABLE:
            try:
                reader_img = PdfReader(pdf_img)
                reader_ocr = PdfReader(BytesIO(pdf_ocr_bytes))
                
                writer = PdfWriter()
                
                if reader_img.pages and reader_ocr.pages:
                    page_img = reader_img.pages[0]
                    page_ocr = reader_ocr.pages[0]
                    
                    if vision_description:
                        vision_pdf = pytesseract.image_to_pdf_or_hocr(
                            Image.new('L', (100, 100), color=255),
                            extension='pdf'
                        )
                        if vision_pdf:
                            reader_vision = PdfReader(BytesIO(vision_pdf))
                            if reader_vision.pages:
                                page_ocr = reader_vision.pages[0]
                    
                    page_img.merge_page(page_ocr)
                    writer.add_page(page_img)
                    
                    output = BytesIO()
                    writer.write(output)
                    return output.getvalue()
            except Exception as e:
                return pdf_ocr_bytes
        
        return pdf_ocr_bytes
    
    except Exception as e:
        pdf_out = BytesIO()
        img.convert('RGB').save(pdf_out, format='PDF')
        return pdf_out.getvalue()

def ocr_pdf(path, vision_description=None):
    """OCR un PDF et retourne (texte, chemin PDF searchable)."""
    try:
        full_text = ""
        page_pdfs = []
        images = []
        
        try:
            with pdfplumber.open(path) as pdf:
                if pdf.pages:
                    for page in pdf.pages:
                        try:
                            img = page.to_image(resolution=300).original
                            if img:
                                images.append(img)
                        except:
                            pass
        except:
            images = []
        
        if (not images or len(images) < 2) and PDF2IMAGE_AVAILABLE:
            try:
                images = convert_from_path(str(path), dpi=300)
            except:
                pass
        
        for img_idx, img in enumerate(images):
            try:
                osd = pytesseract.image_to_osd(img)
                m = re.search(r"Rotate:\s*(\d+)", osd)
                if m:
                    rot = int(m.group(1))
                    if rot:
                        img = img.rotate(360 - rot, expand=True)
            except:
                pass
            
            img = preprocess_image_for_ocr(img)
            text = pytesseract.image_to_string(img, lang="fra")
            full_text += text + "\n"
            
            vision_desc = vision_description if img_idx == 0 and vision_description else None
            if vision_desc:
                print("      └─ Enrichissement vision", flush=True)
            
            pdf_bytes = create_searchable_pdf_page(img, vision_description=vision_desc)
            page_pdfs.append(pdf_bytes)
        
        if not full_text:
            return None, None
        
        with tempfile.NamedTemporaryFile(suffix='.pdf', delete=False) as tmp:
            tmp_path = tmp.name
            _temp_files.append(tmp_path)
            if PYPDF_AVAILABLE and page_pdfs:
                writer = PdfWriter()
                for pb in page_pdfs:
                    if pb:
                        reader = PdfReader(BytesIO(pb))
                        for p in reader.pages:
                            writer.add_page(p)
                with open(tmp_path, 'wb') as f:
                    writer.write(f)
            elif page_pdfs:
                with open(tmp_path, 'wb') as f:
                    f.write(page_pdfs[0])
        
        return full_text, tmp_path
    except:
        return None, None

def extract_from_image(path, vision_description=None):
    """Extrait texte d'une image et génère un PDF searchable."""
    try:
        img = Image.open(path)
        
        try:
            osd = pytesseract.image_to_osd(img)
            m = re.search(r"Rotate:\s*(\d+)", osd)
            if m:
                rot = int(m.group(1))
                if rot:
                    img = img.rotate(360 - rot, expand=True)
        except:
            pass
        
        img = preprocess_image_for_ocr(img)
        text = pytesseract.image_to_string(img, lang="fra")
        
        if vision_description:
            print("      └─ Enrichissement vision", flush=True)
        
        pdf_bytes = create_searchable_pdf_page(img, vision_description=vision_description)
        if pdf_bytes:
            with tempfile.NamedTemporaryFile(suffix='.pdf', delete=False) as tmp:
                tmp.write(pdf_bytes)
                tmp_path = tmp.name
                _temp_files.append(tmp_path)
            return text, tmp_path
        
        return text, None
    except:
        return None, None

def extract_from_docx(path):
    """Extrait texte d'un DOCX."""
    if not DOCX_AVAILABLE:
        return None
    try:
        doc = Document(path)
        return "\n".join([p.text for p in doc.paragraphs])
    except:
        return None

def extract_from_xlsx(path):
    """Extrait texte d'un XLSX."""
    if not XLSX_AVAILABLE:
        return None
    try:
        wb = load_workbook(path)
        text = ""
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            for row in ws.iter_rows(values_only=True):
                text += " ".join([str(v) or "" for v in row]) + "\n"
        return text
    except:
        return None

# ========== EXTRACTION DATES ==========

def extract_dates(text):
    """Extrait la première date valide au format YYYY-MM."""
    if not text:
        return []
    
    current_year = datetime.now().year
    min_year = current_year - 20
    max_year = current_year + 1
    dates = []
    
    # Format 1: Mois nommés (prioritaire)
    month_names = {
        "janvier": "01", "février": "02", "mars": "03", "avril": "04",
        "mai": "05", "juin": "06", "juillet": "07", "août": "08",
        "septembre": "09", "octobre": "10", "novembre": "11", "décembre": "12",
        "january": "01", "february": "02", "march": "03", "april": "04",
        "may": "05", "june": "06", "july": "07", "august": "08",
        "september": "09", "october": "10", "november": "11", "december": "12"
    }
    for month_name, month_num in month_names.items():
        pattern = rf"\b(\d{{1,2}})\s+{month_name}\s+(\d{{4}})\b"
        matches_named = re.findall(pattern, text, re.IGNORECASE)
        for day, year in matches_named:
            day_int = int(day)
            year_int = int(year)
            if 1 <= day_int <= 31 and min_year <= year_int <= max_year:
                dates.append(f"{year}-{month_num}")
                break
        if dates:
            break
    
    # Format 2: YYYY-MM
    if not dates:
        matches_iso = re.findall(r"\b(\d{4})-(\d{2})(?:-\d{2})?\b", text)
        for year, month in matches_iso:
            year_int = int(year)
            month_int = int(month)
            if min_year <= year_int <= max_year and 1 <= month_int <= 12:
                dates.append(f"{year}-{month}")
    
    # Format 3: DD/MM/YYYY
    if not dates:
        matches_slash = re.findall(r"\b(\d{1,2})/(\d{1,2})/(\d{4})\b", text)
        for day, month, year in matches_slash:
            day_int = int(day)
            month_int = int(month)
            year_int = int(year)
            if 1 <= day_int <= 31 and 1 <= month_int <= 12 and min_year <= year_int <= max_year:
                month_str = str(month_int).zfill(2)
                dates.append(f"{year}-{month_str}")
    
    # Format 4: YYYY/MM/DD
    if not dates:
        matches_slash_reverse = re.findall(r"\b(\d{4})/(\d{1,2})/(\d{1,2})\b", text)
        for year, month, day in matches_slash_reverse:
            day_int = int(day)
            month_int = int(month)
            year_int = int(year)
            if 1 <= day_int <= 31 and 1 <= month_int <= 12 and min_year <= year_int <= max_year:
                month_str = str(month_int).zfill(2)
                dates.append(f"{year}-{month_str}")
    
    # Format 5: YYYY seul (fallback)
    if not dates:
        matches_year = re.findall(r"\b(19|20)(\d{2})\b", text)
        for century, year_suffix in matches_year:
            full_year_int = int(century + year_suffix)
            if min_year <= full_year_int <= max_year:
                dates.append(century + year_suffix)
                break
    
    return list(dict.fromkeys(dates[:1]))

# ========== ANALYSE OLLAMA ==========

def extract_first_page(text):
    """Extrait les ~3500 premiers caractères (première page)."""
    lines = text.split('\n')
    page_text = []
    char_count = 0
    
    for line in lines:
        if char_count > 3500:
            break
        page_text.append(line)
        char_count += len(line) + 1
    
    return '\n'.join(page_text)

def analyze_llava(image_path, model=None):
    """Extrait le texte brut visible sur une image via Llava."""
    if model is None:
        model = get_llava_model()
    try:
        with open(image_path, 'rb') as f:
            image_data = base64.b64encode(f.read()).decode('utf-8')
        
        vision_prompt = load_prompt("llava_vision")
        
        response = ollama.generate(
            model=model,
            prompt=vision_prompt,
            images=[image_data],
            stream=False
        )
        
        analysis = response.get("response", "").strip()
        
        # Extraire max 4000 caractères pour avoir tout le texte
        if len(analysis) > 4000:
            analysis = analysis[:4000]
        
        return analysis
    except Exception as e:
        return None


def analyze_ollama(text, dates, model, vision_analysis=None, pass_level="initial", original_filename=None):
    """Analyse le texte avec Ollama et retourne Institution, Objet, Date."""
    dates_str = ", ".join(dates) if dates else "aucune"
    first_page_text = extract_first_page(text)
    
    context_parts = []
    
    if original_filename:
        context_parts.append(f"[NOM FICHIER ORIGINAL]\n{original_filename}")
    
    if vision_analysis:
        context_parts.append(f"[TEXTE LLAVA (VISION IA)]\n{vision_analysis}")
    
    context_parts.append(f"[TEXTE TESSERACT (OCR)]\n{first_page_text}")
    context_parts.append(f"[DATES CANDIDATES]\n{dates_str}")
    
    text_to_send = "\n\n".join(context_parts)
    prompt_template = load_prompt("ollama_analysis", model_name=model)
    prompt = prompt_template.format(dates=dates_str, text=text_to_send)
    
    try:
        response = ollama.generate(model=model, prompt=prompt, stream=False)
        return response.get("response", "").strip()
    except Exception as e:
        print(f"      ❌ Erreur Ollama: {e}")
        return None


def simplify_institution_name(name):
    """Supprime articles/formes juridiques en tête et fin pour garder un nom court."""
    if not name:
        return name
    cleaned = name.strip()
    # Supprimer articles au début
    cleaned = re.sub(r"^(la|le|les|l'|the)\s+", "", cleaned, flags=re.IGNORECASE)
    # Supprimer formes juridiques (avec ou sans points): S.A., S.A.S., SA, SAS, SARL, etc.
    cleaned = re.sub(
        r"\s+(s\.?a\.?(?:s\.?)?|sarl|scs|snc|sca|gmbh|inc\.?|ltd\.?|plc|llc|corp\.?|company|limited|anonyme|soci[eé]t[eé])\s*$",
        "",
        cleaned,
        flags=re.IGNORECASE,
    )
    cleaned = re.sub(r"\s+", " ", cleaned).strip()
    return cleaned or name.strip()


def title_from_first_page(first_page_text):
    """Heuristique : premier intitulé plausible sur la 1ère page."""
    if not first_page_text:
        return None
    for line in first_page_text.splitlines():
        candidate = line.strip()
        if not candidate:
            continue
        if len(candidate) < 6 or len(candidate) > 80:
            continue
        if sum(ch.isalpha() for ch in candidate) < max(5, len(candidate) // 3):
            continue
        if re.match(r"^(page|document|annexe|table(au)?|index|sommaire)", candidate, re.IGNORECASE):
            continue
        return candidate
    return None

def best_match_with_title(variants, title_text):
    """Sélectionne la variante la plus proche du titre du document."""
    if not title_text:
        return variants[0] if variants else "inconnu"
    
    # Normaliser le titre
    title_norm = title_text.lower().strip()
    
    # Scorer chaque variante
    best_variant = variants[0]
    best_score = 0
    
    for variant in variants:
        if variant.lower() == "inconnu":
            continue
        
        variant_norm = variant.lower().strip()
        
        # Score basé sur la similarité avec le titre
        # Plus il y a de mots en commun, meilleur le score
        title_words = set(title_norm.split())
        variant_words = set(variant_norm.split())
        common_words = len(title_words & variant_words)
        
        # Bonus si c'est exactement le titre
        if variant_norm == title_norm:
            score = 1000
        # Bonus si c'est un sous-ensemble du titre
        elif variant_norm in title_norm:
            score = 500 + common_words * 50
        # Sinon, compter les mots en commun
        else:
            score = common_words * 50
        
        if score > best_score:
            best_score = score
            best_variant = variant
    
    return best_variant

def parse_analysis(text, first_page_text=None):
    """Parse la réponse Ollama avec 3 variantes."""
    if not text:
        return None, None, None, False
    
    inst_variants = []
    obj_variants = []
    date = "inconnu"
    
    # Parser strictement chaque ligne
    for line in text.splitlines():
        line_lower = line.lower()
        
        # Institution
        if line_lower.startswith("institution variante"):
            value = line.split(":", 1)[1].strip() if ":" in line else ""
            value = re.sub(r'\s*[\(\[].*$', '', value).strip()
            if value:
                inst_variants.append(value)
        
        # Objet
        elif line_lower.startswith("objet variante"):
            value = line.split(":", 1)[1].strip() if ":" in line else ""
            value = re.sub(r'\s*[\(\[].*$', '', value).strip()
            if value:
                obj_variants.append(value)
        
        # Date
        elif line_lower.startswith("date:"):
            value = line.split(":", 1)[1].strip()
            value = re.sub(r'\s*[\(\[].*$', '', value).strip()
            if re.match(r"^\d{4}-\d{2}$", value):
                date = value
            elif value.lower() != "inconnu":
                match = re.search(r"(\d{4})-(\d{2})", value)
                if match:
                    date = f"{match.group(1)}-{match.group(2)}"
    
    # Assurer au moins 3 variantes (remplir avec "inconnu")
    while len(inst_variants) < 3:
        inst_variants.append("inconnu")
    while len(obj_variants) < 3:
        obj_variants.append("inconnu")
    
    # Sélectionner la meilleure variante basée sur le titre
    title = title_from_first_page(first_page_text)
    inst = best_match_with_title(inst_variants[:3], title)
    obj = best_match_with_title(obj_variants[:3], title)
    
    # Simplifier institution
    inst = simplify_institution_name(inst)
    
    # Validation
    if date == "inconnu" or not re.match(r"^\d{4}-\d{2}$", date):
        return inst, obj, date, False
    
    # Max 1 champ inconnu
    unknown_count = sum(1 for v in [inst, obj] if v == "inconnu")
    certitude = unknown_count <= 1
    return inst, obj, date, certitude

# ========== RENOMMAGE ==========

def sanitize(s):
    """Nettoie agressivement un nom de fichier (no comments)."""
    if not s or s.lower() == "inconnu":
        return "inconnu"
    # Supprimer caractères invalides et commentaires
    s = re.sub(r'[\\/\*?:"<>|\(\)\[\]{}]', '', s)
    # Supprimer caractères spéciaux
    s = re.sub(r'[\n\t\r]', '', s)
    # Nettoyer et retourner
    s = s.strip()
    return s or "inconnu"

def generate_name(inst, obj, date, ext):
    """Génère le nouveau nom - STRICT: YYYY-MM [Institution] [Objet].ext"""
    inst_clean = sanitize(inst)
    obj_clean = sanitize(obj)
    # Format strict: YYYY-MM [Institution] [Objet].ext
    name = f"{date} {inst_clean} {obj_clean}{ext}".strip()
    # Capitaliser chaque mot (sauf la date)
    parts = name.split(' ', 1)  # Séparer date du reste
    if len(parts) == 2:
        return f"{parts[0]} {parts[1].title()}"
    return name.title()

# ========== MAIN ==========

def main():
    check_deps()
    ensure_models()
    config = load_config()
    
    # Dossier source avec défaut
    source_dir = config.get("SOURCE_DIR", "documents")
    choix = input(f"Dossier source [{source_dir}]: ").strip()
    if choix:
        source_dir = choix
    
    if not os.path.isdir(source_dir):
        print("❌ Dossier invalide")
        return
    
    config["SOURCE_DIR"] = source_dir
    save_config(config)
    
    # Sélection modèle (masquer llava qui est utilisé automatiquement pour la vision)
    try:
        out = subprocess.run(['ollama', 'list'], stdout=subprocess.PIPE, text=True, check=True)
        all_models = [l.split()[0] for l in out.stdout.splitlines()[1:] if l.strip()]
        # Filtrer les modèles llava (utilisés automatiquement pour l'analyse vision)
        models = [m for m in all_models if not m.startswith('llava')]
        if models:
            print("\n🧠 Modèles Ollama disponibles:")
            for i, m in enumerate(models, 1):
                print(f"      {i}. {m}")
            choix = input(f"Choix [{config['OLLAMA_MODEL']}]: ").strip()
            if choix.isdigit() and 1 <= int(choix) <= len(models):
                config["OLLAMA_MODEL"] = models[int(choix) - 1]
                save_config(config)
    except:
        pass
    
    # Créer dossiers dans le dossier source avec timestamp
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    export = Path(source_dir) / f"Export_{timestamp}"
    failure = Path(source_dir) / f"Echec_{timestamp}"
    export.mkdir(exist_ok=True)
    failure.mkdir(exist_ok=True)
    
    # Log CSV dans le dossier export
    log_path = export / f"log_{timestamp}.csv"
    with open(log_path, 'w', newline='', encoding='utf-8') as f:
        csv.writer(f).writerow(["Fichier", "Statut", "Nouveau nom", "Institution", "Objet", "Date"])
    
    # Traiter fichiers
    try:
        for file_path in Path(source_dir).iterdir():
            if not file_path.is_file():
                continue
            
            ext = file_path.suffix.lower()
            if ext not in [".pdf", ".png", ".jpg", ".jpeg", ".docx", ".xlsx"]:
                continue
            
            print(f"\n{'='*60}")
            print(f"📄 FICHIER: {file_path.name}")
            print(f"{'='*60}")
            
            # Extraction texte (avec source alternative)
            text_primary = None     # Texte pour analyse initiale
            text_fallback = None    # Texte pour fallback (si certification insuffisante)
            tmp_pdf = None
            
            if ext == ".pdf":
                print("  📖 [PDF] Extraction texte natif...", end=" ", flush=True)
                text_primary = extract_from_pdf(file_path)
                if text_primary:
                    print("✅ (natif)")
                    text_fallback = text_primary  # Source = texte natif
                else:
                    print("❌ (image)")
                    print("  🔍 [OCR] Prétraitement Tesseract...", end=" ", flush=True)
                    text_primary, tmp_pdf = ocr_pdf(file_path)
                    text_fallback = None  # Pas de source alternative pour PDF image
                    if tmp_pdf:
                        print("✅ PDF OCRisé créé")
                    else:
                        print("❌")
            elif ext in [".png", ".jpg", ".jpeg"]:
                print("  🖼️  [IMAGE] OCR Tesseract...", end=" ", flush=True)
                text_primary, tmp_pdf = extract_from_image(file_path)
                text_fallback = None  # Pas de source alternative pour image
                if text_primary:
                    if tmp_pdf:
                        print("✅ PDF OCRisé créé")
                    else:
                        print("✅")
                else:
                    print("❌")
            elif ext == ".docx":
                print("  📝 [DOCX] Extraction texte...", end=" ", flush=True)
                text_primary = extract_from_docx(file_path)
                text_fallback = None
                print("✅" if text_primary else "❌")
            elif ext == ".xlsx":
                print("  📊 [XLSX] Extraction texte...", end=" ", flush=True)
                text_primary = extract_from_xlsx(file_path)
                text_fallback = None
                print("✅" if text_primary else "❌")
            
            if not text_primary:
                print("  ❌ [ERREUR] Aucun texte détecté")
                shutil.copy2(str(file_path), str(failure / file_path.name))
                # Copier aussi le PDF OCRisé s'il existe (pour consultation ultérieure)
                if tmp_pdf and os.path.exists(tmp_pdf):
                    pdf_failure_name = file_path.stem + "_OCR.pdf"
                    shutil.copy2(tmp_pdf, str(failure / pdf_failure_name))
                    print(f"  └─ PDF OCRisé copié: {pdf_failure_name}")
                    _temp_files.remove(tmp_pdf) if tmp_pdf in _temp_files else None
                with open(log_path, 'a', newline='', encoding='utf-8') as f:
                    csv.writer(f).writerow([file_path.name, "Échec", "", "", "", ""])
                continue
            
            # ========== ANALYSE LLAVA (AVANT RECHERCHE DE DATE) ==========
            # Llava analyse la 1ère page et fournit contexte visuel
            vision_analysis = None
            image_for_vision = None
            
            # Déterminer quelle image utiliser pour la vision (PREMIÈRE PAGE UNIQUEMENT)
            if ext in [".png", ".jpg", ".jpeg"]:
                # Image directe = déjà une seule page
                image_for_vision = str(file_path)
            elif ext == ".pdf" and tmp_pdf:
                # Extraire UNIQUEMENT la première page du PDF en image pour llava
                try:
                    images = []
                    with pdfplumber.open(tmp_pdf) as pdf:
                        if pdf.pages:
                            # Convertir PREMIÈRE PAGE uniquement
                            img = pdf.pages[0].to_image(resolution=300).original
                            if img:
                                images.append(img)
                    
                    # Fallback avec pdf2image si besoin (aussi première page uniquement)
                    if not images and PDF2IMAGE_AVAILABLE:
                        images = convert_from_path(tmp_pdf, dpi=300, first_page=1, last_page=1)
                    
                    if images:
                        # Sauvegarder temporairement l'image de la 1ère page
                        with tempfile.NamedTemporaryFile(suffix='.png', delete=False) as tmp_img:
                            images[0].save(tmp_img.name)
                            image_for_vision = tmp_img.name
                            _temp_files.append(tmp_img.name)
                except:
                    image_for_vision = None
            
            # Appeler llava si une image est disponible
            if image_for_vision:
                llava_model = get_llava_model()
                print(f"  👁️  [{llava_model.upper()}] Analyse vision 1ère page...")
                vision_analysis = analyze_llava(image_for_vision)
                if vision_analysis:
                    print(f"      ✅ {len(vision_analysis)} caractères extraits")
                else:
                    print("      ❌ Pas de résultat")
            
            # ========== EXTRACTION DATES (AVEC TESSERACT + LLAVA) ==========
            # Combine Tesseract OCR + description Llava pour meilleure détection
            print("  📅 [DATES] Recherche...")
            # Texte combiné pour recherche de date
            combined_text = text_primary
            if vision_analysis:
                # Ajouter la description llava au texte pour recherche de date
                combined_text = vision_analysis + "\n" + text_primary
            
            dates = extract_dates(combined_text)
            print(f"      ✅ {len(dates)} date(s) trouvée(s): {dates if dates else 'aucune'}")
            
            # RE-GÉNÉRER les PDF OCRisés avec la description vision intégrée
            if vision_analysis:
                print("  🔄 [PDF] Régénération avec enrichissement vision...")
                try:
                    if ext == ".pdf":
                        # Régénérer le PDF avec vision intégrée
                        text_primary, tmp_pdf = ocr_pdf(file_path, vision_description=vision_analysis)
                        print("      ✅ PDF enrichi")
                    elif ext in [".png", ".jpg", ".jpeg"]:
                        # Régénérer le PDF image avec vision intégrée
                        text_primary, tmp_pdf = extract_from_image(file_path, vision_description=vision_analysis)
                        print("      ✅ PDF enrichi")
                except Exception as e:
                    print(f"      ❌ Erreur: {e}")
                    # Continuer avec les PDFs générés précédemment
            
            # Analyse Ollama passe 1 (avec analyse vision optionnelle)
            print("  🧠 [OLLAMA] Analyse IA (passe 1)...")
            analysis = analyze_ollama(text_primary, dates, config["OLLAMA_MODEL"], 
                                      vision_analysis=vision_analysis, pass_level="initial",
                                      original_filename=file_path.name)
            inst, obj, date, certitude = parse_analysis(analysis, extract_first_page(text_primary))
            print(f"      {'✅ Confiance haute' if certitude else '⚠️  Confiance basse'}")
            
            # Passe 2 : si certitude insuffisante ET source alternative disponible
            if not certitude and text_fallback:
                print("  🧠 [OLLAMA] Analyse IA (passe 2 - fallback)...")
                analysis2 = analyze_ollama(text_fallback, dates, config["OLLAMA_MODEL"], 
                                          vision_analysis=vision_analysis, pass_level="fallback",
                                          original_filename=file_path.name)
                inst2, obj2, date2, certitude2 = parse_analysis(analysis2, extract_first_page(text_fallback))
                if certitude2:
                    inst, obj, date, certitude = inst2, obj2, date2, certitude2
                    print("      ✅ Résultats fallback utilisés")
                else:
                    print("      ❌ Passe 1 conservée")
            
            # Validation
            print(f"  🏷️  [RÉSULTAT] {inst} | {obj} | {date}")
            missing_field = any(v == "inconnu" for v in [inst, obj, date])
            date_valid = bool(re.match(r"^\d{4}-\d{2}$", date))
            if missing_field or not date_valid:
                failure_date = date if date_valid else "inconnu"
                failure_name = generate_name(inst or "inconnu", obj or "inconnu", failure_date, ext)
                print("      ❌ ÉCHEC - Champs manquants ou date invalide")
                print(f"      └─ Fichier rejeté: {failure_name}")
                shutil.copy2(str(file_path), str(failure / failure_name))
                # Copier aussi le PDF OCRisé en cas d'échec (pour consultation ultérieure)
                if tmp_pdf and os.path.exists(tmp_pdf):
                    pdf_failure_name = generate_name(inst or "inconnu", obj or "inconnu", failure_date, ".pdf")
                    shutil.copy2(tmp_pdf, str(failure / pdf_failure_name))
                    print(f"  └─ PDF OCRisé copié: {pdf_failure_name}")
                    _temp_files.remove(tmp_pdf) if tmp_pdf in _temp_files else None
                with open(log_path, 'a', newline='', encoding='utf-8') as f:
                    csv.writer(f).writerow([file_path.name, "Échec", failure_name, inst, obj, date])
                continue
            print("      ✅ Validation OK")
            
            # Renommage
            new_name = generate_name(inst, obj, date, ext)
            new_path = export / new_name
            shutil.copy2(str(file_path), str(new_path))
            
            if tmp_pdf:
                pdf_name = generate_name(inst, obj, date, ".pdf")
                shutil.copy2(tmp_pdf, str(export / pdf_name))
                _temp_files.remove(tmp_pdf) if tmp_pdf in _temp_files else None
            
            print(f"  🎉 EXPORTÉ: {new_name}")
            with open(log_path, 'a', newline='', encoding='utf-8') as f:
                csv.writer(f).writerow([file_path.name, "Succès", new_name, inst, obj, date])
    
    except KeyboardInterrupt:
        print("\n\n⚠️  Interrompu par l'utilisateur")
        cleanup_temp_files()
        sys.exit(130)
    except Exception as e:
        print(f"\n❌ Erreur inattendue: {e}")
        cleanup_temp_files()
        raise
    finally:
        cleanup_temp_files()
        print("\n✅ Exécution terminée.")

if __name__ == "__main__":
    main()
