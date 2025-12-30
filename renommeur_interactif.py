#!/usr/bin/env python3
"""
RenAIme Interactif - Script de tri et renommage avec suggestions interactives
Analyse les fichiers en arrière-plan et propose des noms avec choix utilisateur
"""

import os
import sys
import json
import csv
import re
import shutil
import tempfile
import subprocess
import signal
import atexit
import threading
import queue
from pathlib import Path
from datetime import datetime
from collections import deque

import pdfplumber
from PIL import Image, ImageOps, ImageEnhance, ImageFilter
import pytesseract
import ollama

try:
    from docx import Document
    DOCX_AVAILABLE = True
except:
    DOCX_AVAILABLE = False

try:
    from openpyxl import load_workbook
    XLSX_AVAILABLE = True
except:
    XLSX_AVAILABLE = False

try:
    from pypdf import PdfReader, PdfWriter
    PYPDF_AVAILABLE = True
except:
    PYPDF_AVAILABLE = False

try:
    import cv2
    import numpy as np
    CV2_AVAILABLE = True
except:
    CV2_AVAILABLE = False

try:
    from pdf2image import convert_from_path
    PDF2IMAGE_AVAILABLE = True
except:
    PDF2IMAGE_AVAILABLE = False

# ========== CONFIGURATION ==========

DEFAULT_CONFIG = {
    "SOURCE_DIR": "documents",
    "EXPORT_DIR": "Export",
    "FAILURE_DIR": "Echec",
    "OLLAMA_MODEL": "llama3:8b-instruct-q4_0",
}

# Liste globale des fichiers temporaires à nettoyer
_temp_files = []

# Queue pour communication entre threads
analysis_queue = queue.Queue()
user_input_queue = queue.Queue()

class AnalysisResult:
    """Résultat d'analyse avec suggestions."""
    def __init__(self, filename, text=None, tmp_pdf=None):
        self.filename = filename
        self.text = text
        self.tmp_pdf = tmp_pdf
        self.inst = None
        self.obj = None
        self.date = None
        self.suggestions = []
        self.error = None
        self.logs = []  # Logs d'étapes
    
    def log(self, message):
        """Enregistre une étape de traitement."""
        self.logs.append(message)
        print(message, flush=True)

def cleanup_temp_files():
    """Nettoie tous les fichiers temporaires créés lors de l'exécution."""
    global _temp_files
    for tmp in _temp_files[:]:
        try:
            if os.path.exists(tmp):
                os.remove(tmp)
                print(f"  [NETTOYAGE] Fichier temporaire supprimé: {tmp}")
        except Exception as e:
            print(f"  [ERREUR NETTOYAGE] {tmp}: {e}")
        finally:
            _temp_files.remove(tmp) if tmp in _temp_files else None

def signal_handler(signum, frame):
    """Gestionnaire d'interruption (Ctrl+C, SIGTERM)."""
    signal_name = signal.Signals(signum).name
    print(f"\n\n[⚠️  INTERRUPTION] Signal {signal_name} reçu - Nettoyage en cours...")
    cleanup_temp_files()
    print("[✓] Nettoyage terminé. Au revoir!")
    sys.exit(130)

signal.signal(signal.SIGINT, signal_handler)
signal.signal(signal.SIGTERM, signal_handler)
atexit.register(cleanup_temp_files)

def load_config():
    """Charge ou crée la configuration."""
    if os.path.exists("config.json"):
        with open("config.json", "r", encoding="utf-8") as f:
            return {**DEFAULT_CONFIG, **json.load(f)}
    return DEFAULT_CONFIG

def save_config(config):
    """Sauvegarde la configuration."""
    with open("config.json", "w", encoding="utf-8") as f:
        json.dump(config, f, indent=2, ensure_ascii=False)

def check_deps():
    """Vérifie les dépendances."""
    missing = []
    for pkg in ["pdfplumber", "PIL", "pytesseract", "ollama"]:
        try:
            __import__(pkg)
        except:
            missing.append(pkg)
    if missing:
        print(f"[ERREUR] Dépendances manquantes: {', '.join(missing)}")
        sys.exit(1)

# ========== EXTRACTION TEXTE ==========

def preprocess_image_for_ocr(img):
    """Prétraitement OCR - meilleures pratiques pour contraste difficile."""
    img = img.convert('L')
    img = img.filter(ImageFilter.MedianFilter(size=3))
    
    if CV2_AVAILABLE:
        img_cv = cv2.cvtColor(np.array(img), cv2.COLOR_GRAY2BGR) if len(np.array(img).shape) == 2 else np.array(img)
        clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8, 8))
        img_cv_gray = cv2.cvtColor(img_cv, cv2.COLOR_BGR2GRAY)
        img_cv_gray = clahe.apply(img_cv_gray)
        img = Image.fromarray(img_cv_gray)
        
        img = ImageEnhance.Contrast(img).enhance(1.5)
        
        img_cv = cv2.cvtColor(np.array(img), cv2.COLOR_GRAY2BGR) if len(np.array(img).shape) == 2 else np.array(img)
        img_cv_gray = cv2.cvtColor(img_cv, cv2.COLOR_BGR2GRAY) if len(img_cv.shape) == 3 else img_cv
        _, img_cv_gray = cv2.threshold(img_cv_gray, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
        img = Image.fromarray(img_cv_gray)
        
        img_cv = np.array(img)
        kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (2, 2))
        img_cv = cv2.morphologyEx(img_cv, cv2.MORPH_CLOSE, kernel)
        img = Image.fromarray(img_cv)
    else:
        img = ImageEnhance.Contrast(img).enhance(1.5)
    
    img = img.filter(ImageFilter.SHARPEN)
    return img

def create_searchable_pdf_page(img):
    """Crée une page PDF searchable: image visible + couche OCR texte invisible."""
    try:
        from io import BytesIO
        
        pdf_img = BytesIO()
        img_rgb = img.convert('RGB')
        img_rgb.save(pdf_img, format='PDF')
        pdf_img.seek(0)
        
        pdf_ocr_bytes = pytesseract.image_to_pdf_or_hocr(img, extension='pdf')
        if not pdf_ocr_bytes:
            return pdf_img.getvalue()
        
        if PYPDF_AVAILABLE:
            try:
                from io import BytesIO
                reader_img = PdfReader(pdf_img)
                reader_ocr = PdfReader(BytesIO(pdf_ocr_bytes))
                
                writer = PdfWriter()
                
                if reader_img.pages and reader_ocr.pages:
                    page_img = reader_img.pages[0]
                    page_ocr = reader_ocr.pages[0]
                    page_img.merge_page(page_ocr)
                    writer.add_page(page_img)
                    
                    output = BytesIO()
                    writer.write(output)
                    return output.getvalue()
            except Exception as e:
                return pdf_ocr_bytes
        
        return pdf_ocr_bytes
    
    except Exception as e:
        pdf_out = BytesIO()
        img.convert('RGB').save(pdf_out, format='PDF')
        return pdf_out.getvalue()

def extract_from_pdf(path):
    """Extrait texte d'un PDF."""
    try:
        with pdfplumber.open(path) as pdf:
            texts = [page.extract_text() or "" for page in pdf.pages]
        return "\n".join(texts) if any(texts) else None
    except:
        return None

def ocr_pdf(path):
    """OCR un PDF et retourne le texte + chemin temp PDF searchable."""
    try:
        full_text = ""
        page_pdfs = []
        images = []
        
        try:
            with pdfplumber.open(path) as pdf:
                if pdf.pages:
                    for page in pdf.pages:
                        try:
                            img = page.to_image(resolution=300).original
                            if img:
                                images.append(img)
                        except:
                            pass
        except:
            images = []
        
        if (not images or len(images) < 2) and PDF2IMAGE_AVAILABLE:
            try:
                images = convert_from_path(str(path), dpi=300)
            except:
                pass
        
        for img in images:
            try:
                osd = pytesseract.image_to_osd(img)
                m = re.search(r"Rotate:\s*(\d+)", osd)
                if m:
                    rot = int(m.group(1))
                    if rot:
                        img = img.rotate(360 - rot, expand=True)
            except:
                pass
            
            img = preprocess_image_for_ocr(img)
            text = pytesseract.image_to_string(img, lang="fra")
            full_text += text + "\n"
            
            pdf_bytes = create_searchable_pdf_page(img)
            page_pdfs.append(pdf_bytes)
        
        if not full_text:
            return None, None
        
        with tempfile.NamedTemporaryFile(suffix='.pdf', delete=False) as tmp:
            tmp_path = tmp.name
            _temp_files.append(tmp_path)
            if PYPDF_AVAILABLE and page_pdfs:
                writer = PdfWriter()
                for pb in page_pdfs:
                    if pb:
                        from io import BytesIO
                        reader = PdfReader(BytesIO(pb))
                        for p in reader.pages:
                            writer.add_page(p)
                with open(tmp_path, 'wb') as f:
                    writer.write(f)
            elif page_pdfs:
                with open(tmp_path, 'wb') as f:
                    f.write(page_pdfs[0])
        
        return full_text, tmp_path
    except:
        return None, None

def extract_from_image(path):
    """Extrait texte d'une image et génère un PDF searchable."""
    try:
        img = Image.open(path)
        
        try:
            osd = pytesseract.image_to_osd(img)
            m = re.search(r"Rotate:\s*(\d+)", osd)
            if m:
                rot = int(m.group(1))
                if rot:
                    img = img.rotate(360 - rot, expand=True)
        except:
            pass
        
        img = preprocess_image_for_ocr(img)
        text = pytesseract.image_to_string(img, lang="fra")
        
        pdf_bytes = create_searchable_pdf_page(img)
        if pdf_bytes:
            with tempfile.NamedTemporaryFile(suffix='.pdf', delete=False) as tmp:
                tmp.write(pdf_bytes)
                tmp_path = tmp.name
                _temp_files.append(tmp_path)
            return text, tmp_path
        
        return text, None
    except:
        return None, None

def extract_from_docx(path):
    """Extrait texte d'un DOCX."""
    if not DOCX_AVAILABLE:
        return None
    try:
        doc = Document(path)
        return "\n".join([p.text for p in doc.paragraphs])
    except:
        return None

def extract_from_xlsx(path):
    """Extrait texte d'un XLSX."""
    if not XLSX_AVAILABLE:
        return None
    try:
        wb = load_workbook(path)
        text = ""
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            for row in ws.iter_rows(values_only=True):
                text += " ".join([str(v) or "" for v in row]) + "\n"
        return text
    except:
        return None

# ========== EXTRACTION DATES ==========

def extract_dates(text):
    """Extrait dates YYYY-MM."""
    if not text:
        return []
    
    current_year = datetime.now().year
    min_year = current_year - 20
    max_year = current_year + 1
    dates = []
    
    matches_iso = re.findall(r"\b(\d{4})-(\d{2})(?:-\d{2})?\b", text)
    for year, month in matches_iso:
        year_int = int(year)
        month_int = int(month)
        if min_year <= year_int <= max_year and 1 <= month_int <= 12:
            dates.append(f"{year}-{month}")
    
    if not dates:
        matches_year = re.findall(r"\b(19|20)(\d{2})\b", text)
        for century, year_suffix in matches_year:
            full_year_int = int(century + year_suffix)
            if min_year <= full_year_int <= max_year:
                dates.append(century + year_suffix)
                break
    
    if not dates:
        matches_slash = re.findall(r"\b(\d{1,2})/(\d{1,2})/(\d{4})\b", text)
        for day, month, year in matches_slash:
            day_int = int(day)
            month_int = int(month)
            year_int = int(year)
            if 1 <= day_int <= 31 and 1 <= month_int <= 12 and min_year <= year_int <= max_year:
                month_str = str(month_int).zfill(2)
                dates.append(f"{year}-{month_str}")
    
    return list(dict.fromkeys(dates[:1]))

# ========== ANALYSE OLLAMA ==========

def extract_first_page(text):
    """Extrait la première page seulement (~3500 chars pour 4K context window)."""
    lines = text.split('\n')
    page_text = []
    char_count = 0
    
    for line in lines:
        if char_count > 3500:
            break
        page_text.append(line)
        char_count += len(line) + 1
    
    return '\n'.join(page_text)

PROMPT_TEMPLATE = """Tu es un assistant d'analyse de documents. Analyse le texte ci-dessous (première page d'un document) et extrais STRICTEMENT trois champs : Institution, Objet et Date.

IMPORTANT: Pour Institution et Objet, propose 3 variantes différentes classées par confiance (Variante 1 = plus probable, Variante 3 = moins probable).

INSTRUCTIONS DÉTAILLÉES:

1. INSTITUTION (Nom de l'émetteur/organisme)
   - Identifie l'organisation qui émet le document
   - Simplifie AGRESSIVEMENT : supprime articles, formes juridiques
   - Propose 3 variantes différentes (de la plus à la moins probable)
   - Format : Title Case
   - Si impossible, retourne "inconnu"

2. OBJET (Type/Nature du document)
   - Déduis le type GÉNÉRAL du document à partir de son contenu
   - Exemples : "Facture", "Releve Bancaire", "Contrat De Travail", "Fiche De Paie"
   - Propose 3 variantes différentes (de la plus à la moins probable)
   - Format : Title Case, court et descriptif (2-5 mots)
   - Si le type n'est pas identifiable, retourne "inconnu"

3. DATE (Horodatage du document)
   - Cherche la date d'émission du document
   - Format attendu : YYYY-MM (année-mois)
   - Format accepté : YYYY (année seule) en dernier recours
   - Candidates prioritaires : {dates}
   - Si aucune date fiable, retourne "inconnu"

FORMAT DE SORTIE STRICT (chaque ligne sur une nouvelle ligne):
Institution Variante 1: <valeur>
Institution Variante 2: <valeur>
Institution Variante 3: <valeur>
Objet Variante 1: <valeur>
Objet Variante 2: <valeur>
Objet Variante 3: <valeur>
Date: <valeur>

Exemple:
Institution Variante 1: Banque De France
Institution Variante 2: Banque Nationale
Institution Variante 3: inconnu
Objet Variante 1: Releve De Compte
Objet Variante 2: Releve Bancaire
Objet Variante 3: Document Bancaire
Date: 2024-12

Texte du document:
{text}
"""

def analyze_ollama(text, dates, model):
    """Analyse texte avec Ollama."""
    dates_str = ", ".join(dates) if dates else "aucune"
    first_page_text = extract_first_page(text)
    prompt = PROMPT_TEMPLATE.format(dates=dates_str, text=first_page_text)
    
    try:
        response = ollama.generate(model=model, prompt=prompt, stream=False)
        return response.get("response", "").strip()
    except Exception as e:
        print(f"ERREUR Ollama: {e}")
        return None

def simplify_institution_name(name):
    """Supprime articles/formes juridiques en tête et fin pour garder un nom court."""
    if not name:
        return name
    cleaned = name.strip()
    # Supprimer articles au début
    cleaned = re.sub(r"^(la|le|les|l'|the)\s+", "", cleaned, flags=re.IGNORECASE)
    # Supprimer formes juridiques (avec ou sans points): S.A., S.A.S., SA, SAS, SARL, etc.
    cleaned = re.sub(
        r"\s+(s\.?a\.?(?:s\.?)?|sarl|scs|snc|sca|gmbh|inc\.?|ltd\.?|plc|llc|corp\.?|company|limited|anonyme|soci[eé]t[eé])\s*$",
        "",
        cleaned,
        flags=re.IGNORECASE,
    )
    cleaned = re.sub(r"\s+", " ", cleaned).strip()
    return cleaned or name.strip()

def title_from_first_page(first_page_text):
    """Heuristique : premier intitulé plausible sur la 1ère page."""
    if not first_page_text:
        return None
    for line in first_page_text.splitlines():
        candidate = line.strip()
        if not candidate:
            continue
        if len(candidate) < 6 or len(candidate) > 80:
            continue
        if sum(ch.isalpha() for ch in candidate) < max(5, len(candidate) // 3):
            continue
        if re.match(r"^(page|document|annexe|table(au)?|index|sommaire)", candidate, re.IGNORECASE):
            continue
        return candidate
    return None

def best_match_with_title(variants, title_text):
    """Sélectionne la variante la plus proche du titre du document."""
    if not title_text:
        return variants[0] if variants else "inconnu"
    
    # Normaliser le titre
    title_norm = title_text.lower().strip()
    
    # Scorer chaque variante
    best_variant = variants[0]
    best_score = 0
    
    for variant in variants:
        if variant.lower() == "inconnu":
            continue
        
        variant_norm = variant.lower().strip()
        
        # Score basé sur la similarité avec le titre
        # Plus il y a de mots en commun, meilleur le score
        title_words = set(title_norm.split())
        variant_words = set(variant_norm.split())
        common_words = len(title_words & variant_words)
        
        # Bonus si c'est exactement le titre
        if variant_norm == title_norm:
            score = 1000
        # Bonus si c'est un sous-ensemble du titre
        elif variant_norm in title_norm:
            score = 500 + common_words * 50
        # Sinon, compter les mots en commun
        else:
            score = common_words * 50
        
        if score > best_score:
            best_score = score
            best_variant = variant
    
    return best_variant

def parse_analysis(text, first_page_text=None):
    """Parse la réponse Ollama avec 3 variantes."""
    if not text:
        return None, None, None, False
    
    inst_variants = []
    obj_variants = []
    date = "inconnu"
    
    # Parser strictement chaque ligne
    for line in text.splitlines():
        line_lower = line.lower()
        
        # Institution
        if line_lower.startswith("institution variante"):
            value = line.split(":", 1)[1].strip() if ":" in line else ""
            value = re.sub(r'\s*[\(\[].*$', '', value).strip()
            if value:
                inst_variants.append(value)
        
        # Objet
        elif line_lower.startswith("objet variante"):
            value = line.split(":", 1)[1].strip() if ":" in line else ""
            value = re.sub(r'\s*[\(\[].*$', '', value).strip()
            if value:
                obj_variants.append(value)
        
        # Date
        elif line_lower.startswith("date:"):
            value = line.split(":", 1)[1].strip()
            value = re.sub(r'\s*[\(\[].*$', '', value).strip()
            if re.match(r"^\d{4}-\d{2}$", value):
                date = value
            elif value.lower() != "inconnu":
                match = re.search(r"(\d{4})-(\d{2})", value)
                if match:
                    date = f"{match.group(1)}-{match.group(2)}"
    
    # Assurer au moins 3 variantes (remplir avec "inconnu")
    while len(inst_variants) < 3:
        inst_variants.append("inconnu")
    while len(obj_variants) < 3:
        obj_variants.append("inconnu")
    
    # Sélectionner la meilleure variante basée sur le titre
    title = title_from_first_page(first_page_text)
    inst = best_match_with_title(inst_variants[:3], title)
    obj = best_match_with_title(obj_variants[:3], title)
    
    # Simplifier institution
    inst = simplify_institution_name(inst)
    
    # Validation
    if date == "inconnu" or not re.match(r"^\d{4}-\d{2}$", date):
        return inst, obj, date, False
    
    # Max 1 champ inconnu
    unknown_count = sum(1 for v in [inst, obj] if v == "inconnu")
    certitude = unknown_count <= 1
    return inst, obj, date, certitude

# ========== RENOMMAGE & SUGGESTIONS ==========

def sanitize(s):
    """Nettoie agressivement un nom de fichier."""
    if not s or s.lower() == "inconnu":
        return "inconnu"
    s = re.sub(r'[\\/\*?:"<>|\(\)\[\]{}]', '', s)
    s = re.sub(r'[\n\t\r]', '', s)
    s = s.strip()
    return s or "inconnu"

def generate_name(inst, obj, date, ext):
    """Génère le nom principal."""
    inst_clean = sanitize(inst)
    obj_clean = sanitize(obj)
    name = f"{date} {inst_clean} {obj_clean}{ext}".strip()
    parts = name.split(' ', 1)
    if len(parts) == 2:
        return f"{parts[0]} {parts[1].title()}"
    return name.title()

def generate_suggestions(inst, obj, date, ext):
    """Génère 5 suggestions de noms."""
    inst_clean = sanitize(inst)
    obj_clean = sanitize(obj)
    
    suggestions = []
    
    # Suggestion 1: Format standard (YYYY-MM Institution Objet)
    name1 = f"{date} {inst_clean} {obj_clean}{ext}".strip()
    name1 = f"{date} {name1.split(' ', 1)[1].title()}" if ' ' in name1 else name1.title()
    suggestions.append(name1)
    
    # Suggestion 2: Objet en premier (YYYY-MM Objet Institution)
    name2 = f"{date} {obj_clean} {inst_clean}{ext}".strip()
    name2 = f"{date} {name2.split(' ', 1)[1].title()}" if ' ' in name2 else name2.title()
    suggestions.append(name2)
    
    # Suggestion 3: Institution en tête (Institution YYYY-MM Objet)
    name3 = f"{inst_clean} {date} {obj_clean}{ext}".strip()
    name3 = name3.title()
    suggestions.append(name3)
    
    # Suggestion 4: Date à la fin (Institution Objet YYYY-MM)
    name4 = f"{inst_clean} {obj_clean} {date}{ext}".strip()
    name4 = name4.title()
    suggestions.append(name4)
    
    # Suggestion 5: Format compact avec tirets
    name5 = f"{date}-{inst_clean}-{obj_clean}{ext}".strip()
    name5 = name5.replace("_", " ").title().replace(" ", "-")
    suggestions.append(name5)
    
    return suggestions

# ========== THREAD ANALYSE ==========

def analysis_worker(source_dir, model, log_path):
    """Thread worker pour analyser les fichiers en arrière-plan."""
    for file_path in Path(source_dir).iterdir():
        if not file_path.is_file():
            continue
        
        ext = file_path.suffix.lower()
        if ext not in [".pdf", ".png", ".jpg", ".jpeg", ".docx", ".xlsx"]:
            continue
        
        result = AnalysisResult(file_path.name)
        print(f"\n[FILE] {file_path.name}")
        
        # Extraction texte
        text_primary = None
        tmp_pdf = None
        
        if ext == ".pdf":
            print("  [PDF] Extraction texte natif...", end=" ", flush=True)
            text_primary = extract_from_pdf(file_path)
            if text_primary:
                print("✓ (natif)")
            else:
                print("✗ (image)")
                print("  [OCR] Prétraitement...", end=" ", flush=True)
                text_primary, tmp_pdf = ocr_pdf(file_path)
                if tmp_pdf:
                    print(f"✓ (PDF OCRisé créé)")
                else:
                    print("✗")
        elif ext in [".png", ".jpg", ".jpeg"]:
            print("  [OCR] Prétraitement image...", end=" ", flush=True)
            text_primary, tmp_pdf = extract_from_image(file_path)
            if text_primary:
                if tmp_pdf:
                    print(f"✓ (Tesseract FRA + PDF OCRisé créé)")
                else:
                    print("✓ (Tesseract FRA)")
            else:
                print("✗")
        elif ext == ".docx":
            print("  [DOCX] Extraction texte...", end=" ", flush=True)
            text_primary = extract_from_docx(file_path)
            print("✓" if text_primary else "✗")
        elif ext == ".xlsx":
            print("  [XLSX] Extraction texte...", end=" ", flush=True)
            text_primary = extract_from_xlsx(file_path)
            print("✓" if text_primary else "✗")
        
        if not text_primary:
            result.error = "Aucun texte détecté"
            print("  [ERREUR] Aucun texte détecté")
            analysis_queue.put(result)
            continue
        
        result.text = text_primary
        result.tmp_pdf = tmp_pdf
        
        # Extraction dates
        print("  [DATES] Recherche...", end=" ", flush=True)
        dates = extract_dates(text_primary)
        print(f"{len(dates)} trouvée(s)")
        
        # Analyse Ollama
        print("  [OLLAMA] Analyse...", end=" ", flush=True)
        analysis = analyze_ollama(text_primary, dates, model)
        if analysis:
            print("✓")
            first_page_text = extract_first_page(text_primary)
            inst, obj, date, certitude = parse_analysis(analysis, first_page_text)
            result.inst = inst
            result.obj = obj
            result.date = date
            
            # Parsing affichage
            status = "✓ (OK)" if certitude else "✗ (incomplet)"
            print(f"  [PARSE] {inst} | {obj} | {date} {status}")
            
            result.suggestions = generate_suggestions(inst, obj, date, ext)
        else:
            result.error = "Erreur Ollama"
            print("✗")
            print("  [ERREUR] Erreur Ollama")
        
        analysis_queue.put(result)

# ========== MAIN ==========

def main():
    check_deps()
    config = load_config()
    
    source_dir = config.get("SOURCE_DIR", "documents")
    choix = input(f"Dossier source [{source_dir}]: ").strip()
    if choix:
        source_dir = choix
    
    if not os.path.isdir(source_dir):
        print("[ERREUR] Dossier invalide")
        return
    
    config["SOURCE_DIR"] = source_dir
    save_config(config)
    
    # Sélection modèle
    try:
        out = subprocess.run(['ollama', 'list'], stdout=subprocess.PIPE, text=True, check=True)
        models = [l.split()[0] for l in out.stdout.splitlines()[1:] if l.strip()]
        if models:
            print("Modèles Ollama:")
            for i, m in enumerate(models, 1):
                print(f"  {i}. {m}")
            choix = input(f"Choix [{config['OLLAMA_MODEL']}]: ").strip()
            if choix.isdigit() and 1 <= int(choix) <= len(models):
                config["OLLAMA_MODEL"] = models[int(choix) - 1]
                save_config(config)
    except:
        pass
    
    # Créer dossiers
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    export = Path(source_dir) / f"Export_{timestamp}"
    failure = Path(source_dir) / f"Echec_{timestamp}"
    export.mkdir(exist_ok=True)
    failure.mkdir(exist_ok=True)
    
    log_path = export / f"log_{timestamp}.csv"
    with open(log_path, 'w', newline='', encoding='utf-8') as f:
        csv.writer(f).writerow(["Fichier", "Statut", "Nouveau nom", "Institution", "Objet", "Date"])
    
    # Lancer thread d'analyse
    print(f"\n[DÉMARRAGE] Analyse en arrière-plan des fichiers...\n")
    analysis_thread = threading.Thread(target=analysis_worker, args=(source_dir, config["OLLAMA_MODEL"], log_path), daemon=False)
    analysis_thread.start()
    
    # Traiter les résultats au fur et à mesure
    processed = 0
    while True:
        try:
            result = analysis_queue.get(timeout=1)
        except queue.Empty:
            if not analysis_thread.is_alive():
                break
            continue
        
        processed += 1
        print(f"\n[FICHIER {processed}] {result.filename}")
        
        if result.error:
            print(f"  ❌ ERREUR: {result.error}")
            shutil.copy2(str(Path(source_dir) / result.filename), str(failure / result.filename))
            with open(log_path, 'a', newline='', encoding='utf-8') as f:
                csv.writer(f).writerow([result.filename, "Échec", "", "", "", result.error])
            continue
        
        print(f"  Institution: {result.inst}")
        print(f"  Objet: {result.obj}")
        print(f"  Date: {result.date}")
        
        # Afficher suggestions
        print(f"\n  Suggestions de noms:")
        for i, sugg in enumerate(result.suggestions, 1):
            print(f"    {i}. {sugg}")
        
        # Demander choix utilisateur
        while True:
            choix = input(f"\n  Choix (1-5) ou 'q' pour rejeter, 'e' pour personnaliser : ").strip().lower()
            
            if choix == 'q':
                print(f"  ❌ REJETÉ")
                shutil.copy2(str(Path(source_dir) / result.filename), str(failure / result.filename))
                with open(log_path, 'a', newline='', encoding='utf-8') as f:
                    csv.writer(f).writerow([result.filename, "Rejeté", "", result.inst, result.obj, result.date])
                break
            
            elif choix == 'e':
                custom = input("  Nouveau nom: ").strip()
                if custom:
                    new_name = custom
                    print(f"  ✅ EXPORTÉ: {new_name}")
                    shutil.copy2(str(Path(source_dir) / result.filename), str(export / new_name))
                    if result.tmp_pdf:
                        pdf_name = custom.rsplit('.', 1)[0] + '.pdf'
                        shutil.copy2(result.tmp_pdf, str(export / pdf_name))
                        _temp_files.remove(result.tmp_pdf) if result.tmp_pdf in _temp_files else None
                    with open(log_path, 'a', newline='', encoding='utf-8') as f:
                        csv.writer(f).writerow([result.filename, "Succès (personnalisé)", new_name, result.inst, result.obj, result.date])
                    break
            
            elif choix in ['1', '2', '3', '4', '5']:
                idx = int(choix) - 1
                new_name = result.suggestions[idx]
                print(f"  ✅ EXPORTÉ: {new_name}")
                shutil.copy2(str(Path(source_dir) / result.filename), str(export / new_name))
                if result.tmp_pdf:
                    pdf_name = new_name.rsplit('.', 1)[0] + '.pdf'
                    shutil.copy2(result.tmp_pdf, str(export / pdf_name))
                    _temp_files.remove(result.tmp_pdf) if result.tmp_pdf in _temp_files else None
                with open(log_path, 'a', newline='', encoding='utf-8') as f:
                    csv.writer(f).writerow([result.filename, "Succès", new_name, result.inst, result.obj, result.date])
                break
    
    analysis_thread.join()
    cleanup_temp_files()
    print("\n[✓] Exécution terminée.")

if __name__ == "__main__":
    main()
