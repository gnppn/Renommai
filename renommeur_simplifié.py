#!/usr/bin/env python3
"""
RenAIme - Script minimaliste de tri et renommage de documents
Basé sur OCR + Analyse IA Ollama + Vision IA
"""

import os
import sys
import json
import csv
import re
import shutil
import tempfile
import subprocess
import signal
import atexit
import base64
from pathlib import Path
from datetime import datetime
from io import BytesIO

import pdfplumber
from PIL import Image, ImageEnhance, ImageFilter
import pytesseract
import ollama

try:
    from docx import Document
    DOCX_AVAILABLE = True
except:
    DOCX_AVAILABLE = False

try:
    from openpyxl import load_workbook
    XLSX_AVAILABLE = True
except:
    XLSX_AVAILABLE = False

try:
    from pypdf import PdfReader, PdfWriter
    PYPDF_AVAILABLE = True
except:
    PYPDF_AVAILABLE = False

try:
    import cv2
    import numpy as np
    CV2_AVAILABLE = True
except:
    CV2_AVAILABLE = False

try:
    from pdf2image import convert_from_path
    PDF2IMAGE_AVAILABLE = True
except:
    PDF2IMAGE_AVAILABLE = False

# ========== CONFIGURATION ==========

SCRIPT_DIR = Path(__file__).parent
PROMPTS_DIR = SCRIPT_DIR / "prompts"

# Cache pour la taille de contexte des modèles
_model_context_cache = {}

# Modèle Vision sélectionné (llava, minicpm-v, etc.)
_selected_vision_model = None

def get_system_power_level():
    """Détecte la puissance du système et retourne un niveau (low, medium, high).
    
    Basé sur:
    - RAM disponible
    - Présence et VRAM GPU (si nvidia-smi disponible)
    """
    ram_gb = 0
    vram_gb = 0
    
    # Détecter la RAM système
    try:
        with open('/proc/meminfo', 'r') as f:
            for line in f:
                if line.startswith('MemTotal:'):
                    # MemTotal en kB
                    ram_kb = int(line.split()[1])
                    ram_gb = ram_kb / (1024 * 1024)
                    break
    except:
        ram_gb = 8  # Valeur par défaut
    
    # Détecter la VRAM GPU (NVIDIA)
    try:
        result = subprocess.run(
            ['nvidia-smi', '--query-gpu=memory.total', '--format=csv,noheader,nounits'],
            capture_output=True, text=True, timeout=5
        )
        if result.returncode == 0:
            # Prendre la première GPU, valeur en MB
            vram_mb = int(result.stdout.strip().split('\n')[0])
            vram_gb = vram_mb / 1024
    except:
        vram_gb = 0  # Pas de GPU NVIDIA ou nvidia-smi non disponible
    
    # Déterminer le niveau de puissance
    # High: >= 16GB RAM et >= 8GB VRAM, ou >= 32GB RAM
    # Medium: >= 8GB RAM et >= 4GB VRAM, ou >= 16GB RAM
    # Low: reste
    
    if (ram_gb >= 16 and vram_gb >= 8) or ram_gb >= 32 or vram_gb >= 12:
        return 'high', ram_gb, vram_gb
    elif (ram_gb >= 8 and vram_gb >= 4) or ram_gb >= 16 or vram_gb >= 6:
        return 'medium', ram_gb, vram_gb
    else:
        return 'low', ram_gb, vram_gb

def select_vision_model(available_models):
    """Sélectionne le meilleur modèle vision selon la puissance du système.
    
    Modèles vision selon puissance:
    - PC faible (low): minicpm-v - Léger et efficace pour OCR
    - PC moyen/puissant (medium/high): llava-llama3 - Plus performant
    """
    global _selected_vision_model
    
    power_level, ram_gb, vram_gb = get_system_power_level()
    
    # Filtrer les modèles vision disponibles
    vision_models = [m for m in available_models if m.startswith(('llava', 'minicpm-v'))]
    
    # Définir les préférences selon le niveau de puissance
    if power_level == 'low':
        # PC faible: préférer minicpm-v (plus léger)
        preferred = ['minicpm-v:latest', 'minicpm-v', 'llava-llama3:latest', 'llava-llama3']
        default_pull = 'minicpm-v:latest'
    else:
        # PC moyen/puissant: préférer llava-llama3
        preferred = ['llava-llama3:latest', 'llava-llama3', 'minicpm-v:latest', 'minicpm-v']
        default_pull = 'llava-llama3:latest'
    
    # Chercher le meilleur modèle disponible
    for model in preferred:
        if model in vision_models:
            _selected_vision_model = model
            return model, power_level, ram_gb, vram_gb
    
    # Si aucun modèle vision n'est disponible, on retourne celui à télécharger
    _selected_vision_model = default_pull
    return default_pull, power_level, ram_gb, vram_gb

def get_vision_model():
    """Retourne le modèle vision sélectionné."""
    global _selected_vision_model
    return _selected_vision_model or 'llava-llama3:latest'

def get_model_context_size(model_name):
    """Récupère la taille de contexte d'un modèle Ollama."""
    if model_name in _model_context_cache:
        return _model_context_cache[model_name]
    
    try:
        result = subprocess.run(
            ['ollama', 'show', model_name],
            capture_output=True, text=True, timeout=10
        )
        output = result.stdout
        
        # Chercher "context length" ou "num_ctx" dans la sortie
        for line in output.splitlines():
            line_lower = line.lower()
            if 'context' in line_lower or 'num_ctx' in line_lower:
                # Extraire le nombre
                numbers = re.findall(r'\d+', line)
                if numbers:
                    ctx_size = int(numbers[-1])
                    _model_context_cache[model_name] = ctx_size
                    return ctx_size
        
        # Défaut si non trouvé: estimer selon le nom du modèle
        if '32k' in model_name or '32b' in model_name:
            return 32768
        elif '16k' in model_name:
            return 16384
        elif '8k' in model_name or '8b' in model_name:
            return 8192
        else:
            return 4096  # Défaut conservateur
    except:
        return 4096  # Défaut en cas d'erreur

def load_prompt(name, model_name=None):
    """Charge un prompt adapté à la taille du modèle."""
    if model_name:
        ctx_size = get_model_context_size(model_name)
        
        # Sélectionner le prompt selon la taille de contexte
        if ctx_size >= 16384:
            suffix = "_16k"
        elif ctx_size >= 8192:
            suffix = "_8k"
        else:
            suffix = "_4k"
        
        # Essayer le prompt spécifique, sinon le générique
        prompt_file = PROMPTS_DIR / f"{name}{suffix}.txt"
        if prompt_file.exists():
            return prompt_file.read_text(encoding='utf-8')
    
    # Fallback: prompt générique (4k par défaut)
    prompt_file = PROMPTS_DIR / f"{name}.txt"
    if prompt_file.exists():
        return prompt_file.read_text(encoding='utf-8')
    
    # Dernier recours: prompt 4k
    prompt_file = PROMPTS_DIR / f"{name}_4k.txt"
    if prompt_file.exists():
        return prompt_file.read_text(encoding='utf-8')
    
    raise FileNotFoundError(f"Prompt non trouvé: {name}")

DEFAULT_CONFIG = {
    "SOURCE_DIR": "documents",
    "OLLAMA_MODEL": "llama3:8b-instruct-q4_0",
}

# Liste globale des fichiers temporaires à nettoyer
_temp_files = []

def cleanup_temp_files():
    """Nettoie tous les fichiers temporaires créés lors de l'exécution."""
    global _temp_files
    for tmp in _temp_files[:]:
        try:
            if os.path.exists(tmp):
                os.remove(tmp)
        except Exception as e:
            pass
        finally:
            _temp_files.remove(tmp) if tmp in _temp_files else None

def signal_handler(signum, frame):
    """Gestionnaire d'interruption (Ctrl+C, SIGTERM)."""
    signal_name = signal.Signals(signum).name
    print(f"\n\n⚠️  INTERRUPTION ({signal_name}) - Nettoyage en cours...")
    cleanup_temp_files()
    print("✅ Nettoyage terminé. Au revoir!")
    sys.exit(130)

# Enregistrer les gestionnaires de signaux
signal.signal(signal.SIGINT, signal_handler)   # Ctrl+C
signal.signal(signal.SIGTERM, signal_handler)  # Demande d'arrêt système

# Enregistrer la fonction de nettoyage pour appel automatique à la fermeture
atexit.register(cleanup_temp_files)

def load_config():
    """Charge ou crée la configuration."""
    if os.path.exists("config.json"):
        with open("config.json", "r", encoding="utf-8") as f:
            return {**DEFAULT_CONFIG, **json.load(f)}
    return DEFAULT_CONFIG

def save_config(config):
    """Sauvegarde la configuration."""
    with open("config.json", "w", encoding="utf-8") as f:
        json.dump(config, f, indent=2, ensure_ascii=False)

# ========== DÉPENDANCES ==========

def check_deps():
    """Vérifie les dépendances."""
    missing = []
    for pkg in ["pdfplumber", "PIL", "pytesseract", "ollama"]:
        try:
            __import__(pkg)
        except:
            missing.append(pkg)
    if missing:
        print(f"❌ Dépendances manquantes: {', '.join(missing)}")
        sys.exit(1)
    
    # Vérification critique : Tesseract OCR
    try:
        subprocess.run(['tesseract', '--version'], check=True, capture_output=True)
        print("✅ Tesseract OCR installé")
    except (FileNotFoundError, subprocess.CalledProcessError):
        print("❌ ERREUR CRITIQUE: Tesseract OCR non installé")
        print("   Installation requise: https://github.com/tesseract-ocr/tesseract")
        print("   Windows: choco install tesseract ou téléchargement direct")
        sys.exit(1)
    
    # Vérification des langues OCR
    try:
        result = subprocess.run(['tesseract', '--list-langs'], capture_output=True, text=True)
        langs = result.stdout.strip().split('\n')
        if 'fra' not in langs:
            print("⚠️  Langue française (fra) non installée pour OCR")
            print("   Recommandé: télécharger fra.traineddata")
    except:
        print("⚠️  Impossible de vérifier les langues OCR")

def ensure_models():
    """Vérifie et télécharge les modèles Ollama manquants."""
    print("\n🤖 Vérification des modèles Ollama...")
    
    try:
        out = subprocess.run(['ollama', 'list'], stdout=subprocess.PIPE, text=True, check=True)
        available_models = [l.split()[0] for l in out.stdout.splitlines()[1:] if l.strip()]
        
        if available_models:
            print("      ✅ Modèles disponibles")
        else:
            print("      ⚠️  Aucun modèle")
        
        # Sélectionner le modèle vision selon la puissance du PC
        selected_vision, power_level, ram_gb, vram_gb = select_vision_model(available_models)
        
        power_icons = {'high': '🚀', 'medium': '💻', 'low': '📱'}
        power_names = {'high': 'Élevée', 'medium': 'Moyenne', 'low': 'Limitée'}
        print(f"      {power_icons[power_level]} Puissance détectée: {power_names[power_level]} (RAM: {ram_gb:.1f}GB, VRAM: {vram_gb:.1f}GB)")
        print(f"      👁️  Modèle vision sélectionné: {selected_vision}")
        
        # Télécharger le modèle vision si absent
        vision_models = [m for m in available_models if m.startswith(('llava', 'minicpm-v'))]
        text_models = [m for m in available_models if not m.startswith(('llava', 'minicpm-v'))]
        if not vision_models or selected_vision not in vision_models:
            print(f"      ⬇️  Téléchargement {selected_vision}...")
            try:
                subprocess.run(['ollama', 'pull', selected_vision], check=True)
                print(f"      ✅ {selected_vision} téléchargé")
            except subprocess.CalledProcessError as e:
                print(f"      ❌ Échec: {e}")
                print(f"         Téléchargez manuellement: ollama pull {selected_vision}")
        
        # Télécharger llama3 seulement si AUCUN modèle texte n'est présent
        if not text_models:
            print("      ⬇️  Téléchargement llama3:8b-instruct-q4_0...")
            try:
                subprocess.run(['ollama', 'pull', 'llama3:8b-instruct-q4_0'], check=True)
                print("      ✅ llama3:8b-instruct-q4_0 téléchargé")
            except subprocess.CalledProcessError as e:
                print(f"      ❌ Échec: {e}")
                print("         Téléchargez manuellement: ollama pull llama3:8b-instruct-q4_0")
        
        print()
    
    except FileNotFoundError:
        print("      ❌ Ollama non installé")
        print("         https://ollama.ai")
        sys.exit(1)
    except Exception as e:
        print(f"      ⚠️  Erreur: {e}")

# ========== EXTRACTION TEXTE ==========

def preprocess_image_for_ocr(img):
    """Prétraitement image pour améliorer l'OCR."""
    try:
        # Conversion en niveaux de gris
        img = img.convert('L')
        # Filtre médian pour réduire le bruit
        img = img.filter(ImageFilter.MedianFilter(size=3))
        # Léger rehaussement de contraste
        img = ImageEnhance.Contrast(img).enhance(1.2)
        # Léger rehaussement de netteté
        img = img.filter(ImageFilter.SHARPEN)
        return img
    except Exception as e:
        print(f"❌ Erreur prétraitement image: {e}")
        raise

def extract_from_pdf(path):
    """Extrait texte d'un PDF."""
    try:
        with pdfplumber.open(path) as pdf:
            texts = [page.extract_text() or "" for page in pdf.pages]
        result = "\n".join(texts) if any(texts) else None
        if not result:
            print("      ⚠️  PDF sans texte natif (probablement image)")
        return result
    except Exception as e:
        print(f"      ❌ Erreur extraction PDF: {e}")
        return None

def create_searchable_pdf_page(img):
    """Crée une page PDF avec image visible et couche OCR invisible."""
    try:
        pdf_img = BytesIO()
        img_rgb = img.convert('RGB')
        img_rgb.save(pdf_img, format='PDF')
        pdf_img.seek(0)
        
        # Créer la couche OCR avec Tesseract
        pdf_ocr_bytes = pytesseract.image_to_pdf_or_hocr(img, extension='pdf')
        if not pdf_ocr_bytes:
            return pdf_img.getvalue()
        
        if PYPDF_AVAILABLE:
            try:
                reader_img = PdfReader(pdf_img)
                reader_ocr = PdfReader(BytesIO(pdf_ocr_bytes))
                
                writer = PdfWriter()
                
                if reader_img.pages and reader_ocr.pages:
                    page_img = reader_img.pages[0]
                    page_ocr = reader_ocr.pages[0]
                    
                    # Fusionner la couche OCR sur l'image originale
                    page_img.merge_page(page_ocr)
                    writer.add_page(page_img)
                    
                    output = BytesIO()
                    writer.write(output)
                    return output.getvalue()
            except Exception:
                return pdf_ocr_bytes
        
        return pdf_ocr_bytes
    
    except Exception:
        pdf_out = BytesIO()
        img.convert('RGB').save(pdf_out, format='PDF')
        return pdf_out.getvalue()


def _save_debug_images(images, path_stem):
    """Sauvegarde les images de debug pour diagnostic OCR."""
    try:
        debug_dir = Path("debug_failures")
        debug_dir.mkdir(exist_ok=True)
        debug_sub = debug_dir / f"{path_stem}_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        debug_sub.mkdir(exist_ok=True)
        
        images[0].save(debug_sub / 'page1_raw.png')
        
        try:
            osd = pytesseract.image_to_osd(images[0])
            m = re.search(r"Rotate:\s*(\d+)", osd)
            rot = int(m.group(1)) if m else 0
        except Exception:
            rot = 0
        
        corrected = images[0].rotate(360 - rot, expand=True) if rot else images[0]
        corrected.save(debug_sub / 'page1_corrected.png')
        
        try:
            txt = pytesseract.image_to_string(corrected, lang='fra+eng')
            (debug_sub / 'page1_corrected.txt').write_text(txt, encoding='utf-8')
        except Exception:
            pass
    except Exception:
        pass


def _ocr_pdf_fallback(path):
    """Fallback OCR simple pour PDF (utilisé si pypdf non disponible)."""
    try:
        full_text = ""
        images = []
        
        try:
            with pdfplumber.open(path) as pdf:
                for page in pdf.pages:
                    try:
                        img = page.to_image(resolution=300).original
                        if img:
                            images.append(img)
                    except:
                        pass
        except:
            pass
        
        if not images and PDF2IMAGE_AVAILABLE:
            images = convert_from_path(str(path), dpi=300)
        
        if not images:
            return None, None
        
        # OCR et création PDF simple (sans fusion avec original)
        page_pdfs = []
        for img in images:
            # Détecter et corriger la rotation (si présente)
            try:
                osd = pytesseract.image_to_osd(img)
                m = re.search(r"Rotate:\s*(\d+)", osd)
                if m:
                    rot = int(m.group(1))
                    if rot:
                        img = img.rotate(360 - rot, expand=True)
            except Exception:
                pass

            img_proc = preprocess_image_for_ocr(img)
            # Try French then French+English as fallback
            try:
                txt = pytesseract.image_to_string(img_proc, lang="fra")
                if not txt.strip():
                    print("      ⚠️  OCR français vide, essai français+anglais...")
                    txt = pytesseract.image_to_string(img_proc, lang="fra+eng")
                full_text += txt + "\n"
            except Exception as e:
                print(f"      ❌ Erreur OCR Tesseract: {e}")
                print(f"      📄 Page ignorée faute d'OCR")
                continue
            page_pdfs.append(create_searchable_pdf_page(img_proc))

        if not full_text.strip():
            _save_debug_images(images, Path(path).stem)
            return None, None
        
        # Assembler les pages
        with tempfile.NamedTemporaryFile(suffix='.pdf', delete=False) as tmp:
            tmp_path = tmp.name
            _temp_files.append(tmp_path)
            if page_pdfs:
                tmp.write(page_pdfs[0])  # Fallback: première page seulement
        
        return full_text, tmp_path
    except:
        return None, None


def create_searchable_pdf_from_original(original_path):
    """Crée un PDF searchable à partir du fichier original en préservant la qualité.
    
    Préserve l'image originale comme calque visible et ajoute une couche texte OCR.
    
    Args:
        original_path: Chemin vers le fichier original (PDF ou image)
    
    Returns:
        tuple: (texte_ocr, chemin_pdf_searchable) ou (None, None) en cas d'échec
    """
    try:
        ext = Path(original_path).suffix.lower()
        full_text = ""
        
        if ext == ".pdf":
            # PDF original - OCR multi-pages via Tesseract uniquement
            # If pdf2image is not available, use the pdfplumber fallback OCR
            if not PDF2IMAGE_AVAILABLE:
                return _ocr_pdf_fallback(original_path)

            try:
                images = convert_from_path(str(original_path), dpi=300)
            except Exception:
                # Fallback to pdfplumber-based OCR
                return _ocr_pdf_fallback(original_path)
            if not images:
                return _ocr_pdf_fallback(original_path)
            pdf_pages = []
            full_text = ""
            for i, img in enumerate(images, 1):
                # Détecter et corriger la rotation avant OCR
                try:
                    osd = pytesseract.image_to_osd(img)
                    m = re.search(r"Rotate:\s*(\d+)", osd)
                    if m:
                        rot = int(m.group(1))
                        if rot:
                            print(f"      🔄 Rotation page {i}: {rot}°")
                            img = img.rotate(360 - rot, expand=True)
                except Exception as e:
                    print(f"      ⚠️  Impossible détecter rotation page {i}: {e}")

                img_processed = preprocess_image_for_ocr(img)
                try:
                    page_text = pytesseract.image_to_string(img_processed, lang="fra")
                    if not page_text.strip():
                        print(f"      ⚠️  Page {i}: OCR français vide, essai multilingue...")
                        page_text = pytesseract.image_to_string(img_processed, lang="fra+eng")
                    
                    full_text += page_text + "\n"
                    
                    pdf_bytes = pytesseract.image_to_pdf_or_hocr(img_processed, extension='pdf')
                    if pdf_bytes:
                        pdf_pages.append(pdf_bytes)
                    else:
                        print(f"      ❌ Impossible créer PDF searchable page {i}")
                        
                except Exception as e:
                    print(f"      ❌ Erreur OCR page {i}: {e}")
                    print(f"      📄 Page {i} ignorée faute d'OCR")
                    continue
            if not pdf_pages:
                _save_debug_images(images, Path(original_path).stem)
                return None, None
            writer = PdfWriter()
            for page_bytes in pdf_pages:
                reader = PdfReader(BytesIO(page_bytes))
                writer.add_page(reader.pages[0])
            with tempfile.NamedTemporaryFile(suffix='.pdf', delete=False) as tmp:
                writer.write(tmp)
                tmp_path = tmp.name
                _temp_files.append(tmp_path)
            return full_text, tmp_path
        
        elif ext in [".png", ".jpg", ".jpeg"]:
            # Image originale
            img = Image.open(original_path)
            
            # Rotation automatique
            try:
                osd = pytesseract.image_to_osd(img)
                m = re.search(r"Rotate:\s*(\d+)", osd)
                if m:
                    rot = int(m.group(1))
                    if rot:
                        print(f"      🔄 Rotation image: {rot}°")
                        img = img.rotate(360 - rot, expand=True)
            except Exception as e:
                print(f"      ⚠️  Impossible détecter rotation: {e}")
            
            # Créer le PDF à partir de l'image originale (qualité préservée)
            pdf_original = BytesIO()
            img_rgb = img.convert('RGB')
            img_rgb.save(pdf_original, format='PDF', resolution=300)
            pdf_original.seek(0)
            
            # Prétraitement pour OCR
            img_processed = preprocess_image_for_ocr(img)
            try:
                full_text = pytesseract.image_to_string(img_processed, lang="fra")
                if not full_text.strip():
                    print("      ⚠️  OCR français vide, essai multilingue...")
                    full_text = pytesseract.image_to_string(img_processed, lang="fra+eng")
                
                if not full_text.strip():
                    print("      ❌ OCR impossible: aucun texte détecté")
                    return None, None
                    
            except Exception as e:
                print(f"      ❌ Erreur OCR Tesseract: {e}")
                return None, None
            
            # Créer la couche OCR
            try:
                pdf_ocr_bytes = pytesseract.image_to_pdf_or_hocr(img_processed, extension='pdf')
            except Exception as e:
                print(f"      ❌ Erreur création PDF OCR: {e}")
                return full_text, None
            
            if PYPDF_AVAILABLE and pdf_ocr_bytes:
                try:
                    reader_img = PdfReader(pdf_original)
                    reader_ocr = PdfReader(BytesIO(pdf_ocr_bytes))
                    
                    writer = PdfWriter()
                    
                    if reader_img.pages and reader_ocr.pages:
                        page_img = reader_img.pages[0]
                        page_ocr = reader_ocr.pages[0]
                        
                        # Fusionner couche OCR sur l'image originale
                        page_img.merge_page(page_ocr)
                        writer.add_page(page_img)
                        
                        with tempfile.NamedTemporaryFile(suffix='.pdf', delete=False) as tmp:
                            tmp_path = tmp.name
                            _temp_files.append(tmp_path)
                            with open(tmp_path, 'wb') as f:
                                writer.write(f)
                        
                        return full_text, tmp_path
                except:
                    pass
            
            # Fallback: PDF OCR simple
            if pdf_ocr_bytes:
                with tempfile.NamedTemporaryFile(suffix='.pdf', delete=False) as tmp:
                    tmp.write(pdf_ocr_bytes)
                    tmp_path = tmp.name
                    _temp_files.append(tmp_path)
                return full_text, tmp_path
            
            return full_text, None
        
        return None, None
    
    except Exception as e:
        print(f"      ❌ Erreur critique OCR: {e}")
        return None, None


def enrich_pdf_with_vision_text(pdf_path, vision_text, ocr_text=None):
    """Enrichit un PDF searchable avec le texte extrait par le modèle vision.
    
    Ajoute le texte vision dans les métadonnées du PDF (champ Keywords/Subject).
    Cela permet de retrouver le document via une recherche textuelle tout en
    gardant le PDF léger (pas de couche texte supplémentaire).
    
    Args:
        pdf_path: Chemin vers le PDF searchable
        vision_text: Texte extrait par le modèle vision
        ocr_text: Texte OCR (optionnel, pour enrichissement complet)
    
    Returns:
        str: Chemin vers le PDF enrichi (même fichier si modification in-place)
    """
    if not PYPDF_AVAILABLE or not vision_text:
        return pdf_path
    
    try:
        reader = PdfReader(pdf_path)
        writer = PdfWriter()
        
        # Cloner le document complet pour préserver les images et ressources
        writer.append(reader)
        
        # Préparer le texte pour les métadonnées
        # Limiter la taille pour éviter des métadonnées trop volumineuses
        vision_summary = vision_text[:2000] if len(vision_text) > 2000 else vision_text
        
        # Nettoyer le texte pour les métadonnées (enlever les caractères problématiques)
        vision_summary = re.sub(r'[\x00-\x1f\x7f-\x9f]', ' ', vision_summary)
        vision_summary = ' '.join(vision_summary.split())  # Normaliser les espaces
        
        # Ajouter les métadonnées avec le texte vision
        # Le champ "Keywords" est souvent indexé par les moteurs de recherche
        metadata = reader.metadata or {}
        
        # Créer les nouvelles métadonnées
        new_metadata = {
            '/Producer': 'RenAIme - OCR + Vision IA',
            '/Creator': 'RenAIme',
            '/Subject': f'[Vision IA] {vision_summary}',
        }
        
        # Ajouter le texte combiné dans Keywords pour recherche
        keywords_parts = []
        if vision_text:
            keywords_parts.append(vision_summary)
        
        if keywords_parts:
            new_metadata['/Keywords'] = ' | '.join(keywords_parts)
        
        # Appliquer les métadonnées
        writer.add_metadata(new_metadata)
        
        # Sauvegarder le PDF enrichi
        with tempfile.NamedTemporaryFile(suffix='.pdf', delete=False) as tmp:
            enriched_path = tmp.name
            _temp_files.append(enriched_path)
        
        with open(enriched_path, 'wb') as f:
            writer.write(f)
        
        return enriched_path
    
    except Exception as e:
        # En cas d'erreur, retourner le PDF original
        return pdf_path


def extract_from_docx(path):
    """Extrait texte d'un DOCX."""
    if not DOCX_AVAILABLE:
        return None
    try:
        doc = Document(path)
        return "\n".join([p.text for p in doc.paragraphs])
    except:
        return None

def extract_from_xlsx(path):
    """Extrait texte d'un XLSX."""
    if not XLSX_AVAILABLE:
        return None
    try:
        wb = load_workbook(path)
        text = ""
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            for row in ws.iter_rows(values_only=True):
                text += " ".join([str(v) or "" for v in row]) + "\n"
        return text
    except:
        return None

# ========== EXTRACTION DATES ==========

def _year_2_to_4(year_2):
    """Convertit une année sur 2 chiffres en 4 chiffres."""
    year_int = int(year_2)
    return f"20{year_2}" if year_int < 50 else f"19{year_2}"

def extract_dates(text):
    """Extrait les dates valides au format YYYY-MM (max 5 candidates)."""
    if not text:
        return []
    
    current_year = datetime.now().year
    min_year = current_year - 20
    max_year = current_year + 1
    dates = []
    
    # Dictionnaire des mois (complet + abrégés)
    month_names = {
        # Français complet
        "janvier": "01", "février": "02", "fevrier": "02", "mars": "03", "avril": "04",
        "mai": "05", "juin": "06", "juillet": "07", "août": "08", "aout": "08",
        "septembre": "09", "octobre": "10", "novembre": "11", "décembre": "12", "decembre": "12",
        # Français abrégé
        "janv": "01", "jan": "01", "fév": "02", "fev": "02", "févr": "02", "fevr": "02",
        "avr": "04", "juil": "07", "juill": "07", "sept": "09", "oct": "10", "nov": "11", "déc": "12", "dec": "12",
        # Anglais complet
        "january": "01", "february": "02", "march": "03", "april": "04",
        "may": "05", "june": "06", "july": "07", "august": "08",
        "september": "09", "october": "10", "november": "11", "december": "12",
        # Anglais abrégé
        "feb": "02", "mar": "03", "apr": "04", "jun": "06", "jul": "07", "aug": "08", "sep": "09"
    }
    
    def add_date(year, month):
        """Ajoute une date si valide et non dupliquée."""
        year_int = int(year)
        month_int = int(month)
        if min_year <= year_int <= max_year and 1 <= month_int <= 12:
            date_str = f"{year}-{str(month_int).zfill(2)}"
            if date_str not in dates:
                dates.append(date_str)
    
    # Format 1: "JJ mois YYYY" ou "JJ mois. YYYY" (ex: "15 mars 2024", "15 janv. 2024")
    for month_name, month_num in month_names.items():
        pattern = rf"\b(\d{{1,2}})\s+{month_name}\.?\s+(\d{{4}})\b"
        for day, year in re.findall(pattern, text, re.IGNORECASE):
            if 1 <= int(day) <= 31:
                add_date(year, month_num)
    
    # Format 2: "mois YYYY" sans jour (ex: "mars 2024", "décembre 2023")
    for month_name, month_num in month_names.items():
        pattern = rf"\b{month_name}\.?\s+(\d{{4}})\b"
        for year in re.findall(pattern, text, re.IGNORECASE):
            add_date(year, month_num)
    
    # Format 3: YYYY-MM ou YYYY-MM-DD (ISO)
    for year, month in re.findall(r"\b(\d{4})-(\d{2})(?:-\d{2})?\b", text):
        add_date(year, month)
    
    # Formats DD/MM/YYYY, DD.MM.YYYY, DD-MM-YYYY (séparateurs: / . -)
    for sep in [r"/", r"\.", r"-"]:
        # 4 chiffres pour l'année
        for day, month, year in re.findall(rf"\b(\d{{1,2}}){sep}(\d{{1,2}}){sep}(\d{{4}})\b", text):
            if 1 <= int(day) <= 31:
                add_date(year, month)
        # 2 chiffres pour l'année
        for day, month, year in re.findall(rf"\b(\d{{1,2}}){sep}(\d{{1,2}}){sep}(\d{{2}})\b", text):
            if 1 <= int(day) <= 31:
                add_date(_year_2_to_4(year), month)
    
    # Format YYYY/MM/DD
    for year, month, day in re.findall(r"\b(\d{4})/(\d{1,2})/(\d{1,2})\b", text):
        if 1 <= int(day) <= 31:
            add_date(year, month)
    
    # Format YYYY seul (fallback, moins précis)
    if not dates:
        for century, year_suffix in re.findall(r"\b(19|20)(\d{2})\b", text):
            full_year = century + year_suffix
            if min_year <= int(full_year) <= max_year and full_year not in dates:
                dates.append(full_year)
    
    return dates[:5]  # Maximum 5 dates candidates

# ========== ANALYSE OLLAMA ==========

def extract_first_page(text):
    """Extrait les ~3500 premiers caractères (première page)."""
    lines = text.split('\n')
    page_text = []
    char_count = 0
    
    for line in lines:
        if char_count > 3500:
            break
        page_text.append(line)
        char_count += len(line) + 1
    
    return '\n'.join(page_text)

def analyze_vision(image_path, model=None):
    """Extrait le texte brut ET les informations structurées d'une image via le modèle vision."""
    if model is None:
        model = get_vision_model()
    try:
        with open(image_path, 'rb') as f:
            image_data = base64.b64encode(f.read()).decode('utf-8')

        vision_prompt = load_prompt("vision_prompt_simplifié")

        response = ollama.generate(
            model=model,
            prompt=vision_prompt,
            images=[image_data],
            stream=False
        )

        full_response = response.get("response", "").strip()

        # Parser la réponse pour séparer texte et analyse structurée
        text_content = ""
        structured_content = ""

        if "[TEXT]" in full_response and "[/TEXT]" in full_response:
            # Extraire le contenu entre [TEXT] et [/TEXT]
            text_start = full_response.find("[TEXT]") + len("[TEXT]")
            text_end = full_response.find("[/TEXT]")
            text_content = full_response[text_start:text_end].strip()

            # Le reste après [/TEXT] contient l'analyse structurée
            structured_content = full_response[text_end + len("[/TEXT]"):].strip()
        else:
            # Erreur de format: les balises attendues ne sont pas présentes
            print(f"      ❌ Erreur parsing vision: balises [TEXT] manquantes")
            return None

        # Vérifier que les deux parties sont présentes
        if not text_content:
            print(f"      ❌ Erreur parsing vision: contenu TEXT vide")
            return None

        if not structured_content:
            print(f"      ❌ Erreur parsing vision: contenu structuré vide")
            return None

        return {
            "text": text_content,
            "structured": structured_content
        }
    except Exception as e:
        return None


def analyze_ollama(text, dates, model, vision_analysis=None, pass_level="initial", original_filename=None):
    """Analyse le texte avec Ollama et retourne Institution, Objet, Date."""
    dates_str = ", ".join(dates) if dates else "aucune"
    first_page_text = extract_first_page(text)
    
    context_parts = []
    
    if original_filename:
        context_parts.append(f"[NOM FICHIER ORIGINAL]\n{original_filename}")
    
    if vision_analysis:
        # Accept either a dict returned by analyze_vision or a plain string
        if isinstance(vision_analysis, dict):
            vision_text_for_context = vision_analysis.get("text", "")
        else:
            vision_text_for_context = vision_analysis
        if vision_text_for_context:
            context_parts.append(f"[TEXTE VISION IA]\n{vision_text_for_context}")

    context_parts.append(f"[TEXTE TESSERACT (OCR)]\n{first_page_text}")
    context_parts.append(f"[DATES CANDIDATES]\n{dates_str}")
    
    text_to_send = "\n\n".join(context_parts)
    prompt_template = load_prompt("ollama_analysis", model_name=model)
    prompt = prompt_template.format(dates=dates_str, text=text_to_send)
    
    try:
        response = ollama.generate(model=model, prompt=prompt, stream=False)
        result = response.get("response", "").strip()
        # Debug: afficher la réponse brute
        print(f"      📝 Réponse IA: {result[:200]}..." if len(result) > 200 else f"      📝 Réponse IA: {result}")
        return result
    except Exception as e:
        print(f"      ❌ Erreur Ollama: {e}")
        return None


def simplify_institution_name(name):
    """Supprime articles/formes juridiques en tête et fin pour garder un nom court."""
    if not name:
        return name
    cleaned = name.strip()
    # Supprimer formes juridiques (avec ou sans points): S.A., S.A.S., SA, SAS, SARL, etc.
    cleaned = re.sub(
        r"\s+(s\.?a\.?(:?s\.?)?|sarl|scs|snc|sca|gmbh|inc\.?|ltd\.?|plc|llc|corp\.?|company|limited|anonyme|soci[eé]t[eé])\s*$",
        "",
        cleaned,
        flags=re.IGNORECASE,
    )
    cleaned = re.sub(r"\s+", " ", cleaned).strip()
    return cleaned or name.strip()


def title_from_first_page(first_page_text):
    """Heuristique : premier intitulé plausible sur la 1ère page."""
    if not first_page_text:
        return None
    for line in first_page_text.splitlines():
        candidate = line.strip()
        if not candidate:
            continue
        if len(candidate) < 6 or len(candidate) > 80:
            continue
        if sum(ch.isalpha() for ch in candidate) < max(5, len(candidate) // 3):
            continue
        if re.match(r"^(page|document|annexe|table(au)?|index|sommaire)", candidate, re.IGNORECASE):
            continue
        return candidate
    return None

def best_match_with_title(variants, title_text):
    """Sélectionne la variante la plus proche du titre du document."""
    if not title_text:
        return variants[0] if variants else "inconnu"
    
    # Normaliser le titre
    title_norm = title_text.lower().strip()
    
    # Scorer chaque variante
    best_variant = variants[0]
    best_score = 0
    
    for variant in variants:
        if variant.lower() == "inconnu":
            continue
        
        variant_norm = variant.lower().strip()
        
        # Score basé sur la similarité avec le titre
        # Plus il y a de mots en commun, meilleur le score
        title_words = set(title_norm.split())
        variant_words = set(variant_norm.split())
        common_words = len(title_words & variant_words)
        
        # Bonus si c'est exactement le titre
        if variant_norm == title_norm:
            score = 1000
        # Bonus si c'est un sous-ensemble du titre
        elif variant_norm in title_norm:
            score = 500 + common_words * 50
        # Sinon, compter les mots en commun
        else:
            score = common_words * 50
        
        if score > best_score:
            best_score = score
            best_variant = variant
    
    return best_variant

def parse_analysis(text, first_page_text=None):
    """Parse la réponse Ollama avec 3 variantes."""
    if not text:
        return None, None, None, False
    
    inst_variants = []
    obj_variants = []
    date = "inconnu"
    
    # Parser chaque ligne en nettoyant les préfixes markdown/tirets
    for line in text.splitlines():
        # Nettoyer la ligne: supprimer **, -, →, *, etc. au début
        clean_line = re.sub(r'^[\s\*\-→•]+', '', line).strip()
        # Supprimer aussi les ** autour des mots clés
        clean_line = clean_line.replace('**', '')
        line_lower = clean_line.lower()
        
        # Institution - accepter "Institution 1:", "Institution variante 1:", "Institution:", etc.
        if line_lower.startswith("institution"):
            # Ignorer les lignes qui ne contiennent pas de valeur (juste un numéro)
            if ":" in clean_line:
                value = clean_line.split(":", 1)[1].strip()
                value = re.sub(r'\s*[\(\[].*$', '', value).strip()
                if value and value.lower() not in ("", "inconnu"):
                    inst_variants.append(value)
        
        # Objet - accepter "Objet 1:", "Object 1:", etc.
        elif line_lower.startswith("objet") or line_lower.startswith("object"):
            if ":" in clean_line:
                value = clean_line.split(":", 1)[1].strip()
                value = re.sub(r'\s*[\(\[].*$', '', value).strip()
                if value and value.lower() not in ("", "inconnu"):
                    obj_variants.append(value)
        
        # Date - accepter plusieurs formats
        elif line_lower.startswith("date"):
            # Extraire après le premier : s'il existe
            if ":" in clean_line:
                value = clean_line.split(":", 1)[1].strip()
            else:
                value = clean_line.strip()
            value = re.sub(r'\s*[\(\[].*$', '', value).strip()
            # Format YYYY-MM direct
            if re.match(r"^\d{4}-\d{2}$", value):
                date = value
            elif re.match(r"^\d{4}$", value):
                # Si année seule, compléter avec -01
                date = f"{value}-01"
            elif value.lower() != "inconnu":
                # Format YYYY-MM dans le texte
                match = re.search(r"(\d{4})-(\d{2})", value)
                if match:
                    date = f"{match.group(1)}-{match.group(2)}"
                else:
                    # Format DD/MM/YY ou DD/MM/YYYY - TESTER AVANT année seule
                    match = re.search(r"(\d{1,2})[/\.-](\d{1,2})[/\.-](\d{2,4})", value)
                    if match:
                        day, month, year = match.groups()
                        # Convertir année 2 chiffres en 4 chiffres
                        if len(year) == 2:
                            year = "20" + year if int(year) < 50 else "19" + year
                        date = f"{year}-{month.zfill(2)}"
                    else:
                        # Format YYYY seul (fallback)
                        match = re.search(r"(\d{4})", value)
                        if match:
                            date = f"{match.group(1)}-01"
    
    # Assurer au moins 3 variantes (remplir avec "inconnu")
    while len(inst_variants) < 3:
        inst_variants.append("inconnu")
    while len(obj_variants) < 3:
        obj_variants.append("inconnu")
    
    # Sélectionner la meilleure variante basée sur le titre
    title = title_from_first_page(first_page_text)
    inst = best_match_with_title(inst_variants[:3], title)
    obj = best_match_with_title(obj_variants[:3], title)
    
    # Simplifier institution
    inst = simplify_institution_name(inst)
    
    # Validation
    if date == "inconnu" or not re.match(r"^\d{4}-\d{2}$", date):
        return inst, obj, date, False
    
    # Max 1 champ inconnu
    unknown_count = sum(1 for v in [inst, obj] if v == "inconnu")
    certitude = unknown_count <= 1
    return inst, obj, date, certitude

# ========== RENOMMAGE ==========

def sanitize(s):
    """Nettoie agressivement un nom de fichier (no comments)."""
    if not s or s.lower() == "inconnu":
        return "inconnu"
    # Supprimer caractères invalides et commentaires
    s = re.sub(r'[\\/\*?:"<>|\(\)\[\]{}]', '', s)
    # Supprimer caractères spéciaux
    s = re.sub(r'[\n\t\r]', '', s)
    # Nettoyer et retourner
    s = s.strip()
    return s or "inconnu"

def generate_name(inst, obj, date, ext):
    """Génère le nouveau nom - STRICT: YYYY-MM [Institution] [Objet].ext"""
    inst_clean = sanitize(inst)
    obj_clean = sanitize(obj)
    
    # Capitaliser uniquement la première lettre de chaque partie (sauf date et extension)
    def capitalize_part(s):
        """Première lettre majuscule, reste minuscule pour chaque mot."""
        if not s or s.lower() == "inconnu":
            return s
        # Capitaliser chaque mot individuellement
        words = s.split()
        return ' '.join(w.capitalize() for w in words)
    
    inst_final = capitalize_part(inst_clean)
    obj_final = capitalize_part(obj_clean)
    
    # Format strict: YYYY-MM [Institution] [Objet].ext (extension en minuscules)
    return f"{date} {inst_final} {obj_final}{ext.lower()}"

# ========== MAIN ==========

def main():
    check_deps()
    ensure_models()
    config = load_config()
    
    # Dossier source avec défaut
    source_dir = config.get("SOURCE_DIR", "documents")
    choix = input(f"Dossier source [{source_dir}]: ").strip()
    if choix:
        source_dir = choix
    
    if not os.path.isdir(source_dir):
        print("❌ Dossier invalide")
        return
    
    config["SOURCE_DIR"] = source_dir
    save_config(config)
    
    # Sélection modèle (masquer les modèles vision utilisés automatiquement)
    try:
        out = subprocess.run(['ollama', 'list'], stdout=subprocess.PIPE, text=True, check=True)
        all_models = [l.split()[0] for l in out.stdout.splitlines()[1:] if l.strip()]
        # Filtrer les modèles vision (utilisés automatiquement pour l'analyse vision)
        models = [m for m in all_models if not m.startswith(('llava', 'minicpm-v'))]
        if models:
            print("\n🧠 Modèles Ollama disponibles:")
            for i, m in enumerate(models, 1):
                print(f"      {i}. {m}")
            choix = input(f"Choix [{config['OLLAMA_MODEL']}]: ").strip()
            if choix.isdigit() and 1 <= int(choix) <= len(models):
                config["OLLAMA_MODEL"] = models[int(choix) - 1]
                save_config(config)
    except:
        pass
    
    # Créer dossiers dans le dossier source avec timestamp
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    export = Path(source_dir) / f"Export_{timestamp}"
    failure = Path(source_dir) / f"Echec_{timestamp}"
    export.mkdir(exist_ok=True)
    failure.mkdir(exist_ok=True)
    
    # Log CSV dans le dossier export
    log_path = export / f"log_{timestamp}.csv"
    with open(log_path, 'w', newline='', encoding='utf-8') as f:
        csv.writer(f).writerow(["Fichier", "Statut", "Nouveau nom", "Institution", "Objet", "Date"])
    
    # Traiter fichiers
    try:
        for file_path in Path(source_dir).iterdir():
            if not file_path.is_file():
                continue
            
            ext = file_path.suffix.lower()
            if ext not in [".pdf", ".png", ".jpg", ".jpeg", ".docx", ".xlsx"]:
                continue
            
            print(f"\n{'='*60}")
            print(f"📄 FICHIER: {file_path.name}")
            print(f"{'='*60}")
            
            # Extraction texte (avec source alternative)
            text_primary = None     # Texte pour analyse initiale
            text_fallback = None    # Texte pour fallback (si certification insuffisante)
            tmp_pdf = None
            
            if ext == ".pdf":
                print("  📖 [PDF] Extraction texte natif...", end=" ", flush=True)
                text_primary = extract_from_pdf(file_path)
                if text_primary:
                    print("✅ (natif)")
                    text_fallback = text_primary  # Source = texte natif
                    # Pas besoin d'OCRisation pour PDF avec texte natif
                else:
                    print("❌ (image)")
                    print("  🔍 [OCR] Création PDF searchable (original + couche texte)...", end=" ", flush=True)
                    text_primary, tmp_pdf = create_searchable_pdf_from_original(file_path)
                    text_fallback = None  # Pas de source alternative pour PDF image
                    if tmp_pdf:
                        print("✅ PDF searchable créé")
                    else:
                        print("❌")
            elif ext in [".png", ".jpg", ".jpeg"]:
                print("  🖼️  [IMAGE] Création PDF searchable (original + couche texte)...", end=" ", flush=True)
                text_primary, tmp_pdf = create_searchable_pdf_from_original(file_path)
                text_fallback = None  # Pas de source alternative pour image
                if text_primary:
                    if tmp_pdf:
                        print("✅ PDF searchable créé")
                    else:
                        print("✅")
                else:
                    print("❌")
            elif ext == ".docx":
                print("  📝 [DOCX] Extraction texte...", end=" ", flush=True)
                text_primary = extract_from_docx(file_path)
                text_fallback = None
                print("✅" if text_primary else "❌")
            elif ext == ".xlsx":
                print("  📊 [XLSX] Extraction texte...", end=" ", flush=True)
                text_primary = extract_from_xlsx(file_path)
                text_fallback = None
                print("✅" if text_primary else "❌")
            
            if not text_primary:
                print("  ❌ [ERREUR OCR OBLIGATOIRE] Aucun texte détecté")
                print("      📋 OCR est requis pour tous les documents")
                print("      🔧 Vérifiez : qualité du document, installation Tesseract, langues OCR")
                shutil.copy2(str(file_path), str(failure / file_path.name))
                # Copier aussi le PDF OCRisé s'il existe (pour consultation ultérieure)
                if tmp_pdf and os.path.exists(tmp_pdf):
                    pdf_failure_name = file_path.stem + "_OCR.pdf"
                    shutil.copy2(tmp_pdf, str(failure / pdf_failure_name))
                    print(f"  └─ PDF OCRisé copié: {pdf_failure_name}")
                    _temp_files.remove(tmp_pdf) if tmp_pdf in _temp_files else None
                with open(log_path, 'a', newline='', encoding='utf-8') as f:
                    csv.writer(f).writerow([file_path.name, "Échec OCR", "", "OCR_OBLIGATOIRE", "TEXTE_NON_DETECTE", ""])
                continue
            
            # ========== ANALYSE VISION IA (AVANT RECHERCHE DE DATE) ==========
            # Le modèle vision analyse la 1ère page et fournit contexte visuel
            vision_analysis = None
            image_for_vision = None
            
            # Déterminer quelle image utiliser pour la vision (PREMIÈRE PAGE UNIQUEMENT)
            if ext in [".png", ".jpg", ".jpeg"]:
                # Image directe = déjà une seule page
                image_for_vision = str(file_path)
            elif ext == ".pdf" and tmp_pdf:
                # Extraire UNIQUEMENT la première page du PDF en image pour le modèle vision
                try:
                    images = []
                    with pdfplumber.open(tmp_pdf) as pdf:
                        if pdf.pages:
                            # Convertir PREMIÈRE PAGE uniquement
                            img = pdf.pages[0].to_image(resolution=300).original
                            if img:
                                images.append(img)
                    
                    # Fallback avec pdf2image si besoin (aussi première page uniquement)
                    if not images and PDF2IMAGE_AVAILABLE:
                        images = convert_from_path(tmp_pdf, dpi=300, first_page=1, last_page=1)
                    
                    if images:
                        # Sauvegarder temporairement l'image de la 1ère page
                        with tempfile.NamedTemporaryFile(suffix='.png', delete=False) as tmp_img:
                            images[0].save(tmp_img.name)
                            image_for_vision = tmp_img.name
                            _temp_files.append(tmp_img.name)
                except:
                    image_for_vision = None
            
            # Appeler le modèle vision si une image est disponible
            if image_for_vision:
                vision_model = get_vision_model()
                print(f"  👁️  [{vision_model.upper()}] Analyse vision 1ère page...")
                vision_analysis = analyze_vision(image_for_vision)
                if vision_analysis and vision_analysis.get("text"):
                    text_len = len(vision_analysis.get("text", ""))
                    print(f"      ✅ {text_len} caractères texte + analyse structurée")
                else:
                    print("      ❌ Pas de résultat")
            
            # ========== EXTRACTION DATES (AVEC TESSERACT + VISION IA) ==========
            # Combine Tesseract OCR + description Vision IA pour meilleure détection
            print("  📅 [DATES] Recherche...")
            # Texte combiné pour recherche de date
            combined_text = text_primary
            vision_text = ""
            vision_structured = ""

            if vision_analysis and isinstance(vision_analysis, dict):
                vision_text = vision_analysis.get("text", "")
                vision_structured = vision_analysis.get("structured", "")
                if vision_text:
                    combined_text = vision_text + "\n" + text_primary

            dates = extract_dates(combined_text)
            print(f"      ✅ {len(dates)} date(s) trouvée(s): {dates if dates else 'aucune'}")

            # Analyse structurée : utiliser d'abord la vision, puis Ollama en fallback
            inst, obj, date, certitude = None, None, None, False

            if vision_structured:
                print("  🧠 [VISION] Analyse structurée...")
                inst, obj, date, certitude = parse_analysis(vision_structured, extract_first_page(text_primary))
                print(f"      {'✅ Confiance haute' if certitude else '⚠️  Confiance basse'} (vision)")

            # Fallback vers Ollama si vision insuffisante ou absente
            if not certitude:
                print("  🧠 [OLLAMA] Analyse IA (fallback)...")
                analysis = analyze_ollama(text_primary, dates, config["OLLAMA_MODEL"],
                                          vision_analysis=vision_text, pass_level="initial",
                                          original_filename=file_path.name)
                inst2, obj2, date2, certitude2 = parse_analysis(analysis, extract_first_page(text_primary))
                if certitude2 or not certitude:  # Utiliser Ollama si meilleure ou si vision a échoué
                    inst, obj, date, certitude = inst2, obj2, date2, certitude2
                    print(f"      {'✅ Confiance haute' if certitude else '⚠️  Confiance basse'} (ollama)")

            # Passe 2 : si certitude insuffisante ET source alternative disponible
            if not certitude and text_fallback:
                print("  🧠 [OLLAMA] Analyse IA (passe 2 - fallback)...")
                analysis2 = analyze_ollama(text_fallback, dates, config["OLLAMA_MODEL"],
                                          vision_analysis=vision_text, pass_level="fallback",
                                          original_filename=file_path.name)
                inst2, obj2, date2, certitude2 = parse_analysis(analysis2, extract_first_page(text_fallback))
                if certitude2:
                    inst, obj, date, certitude = inst2, obj2, date2, certitude2
                    print("      ✅ Résultats fallback utilisés")
                else:
                    print("      ❌ Passe 1 conservée")
            
            # Validation
            print(f"  🏷️  [RÉSULTAT] {inst} | {obj} | {date}")
            missing_field = any(v == "inconnu" for v in [inst, obj, date])
            date_valid = bool(re.match(r"^\d{4}-\d{2}$", date))
            if missing_field or not date_valid:
                failure_date = date if date_valid else "inconnu"
                failure_name = generate_name(inst or "inconnu", obj or "inconnu", failure_date, ext)
                print("      ❌ ÉCHEC - Champs manquants ou date invalide")
                print(f"      └─ Fichier rejeté: {failure_name}")
                target_path = failure / failure_name
                try:
                    # Attempt copy; handle unexpected filename characters
                    shutil.copy2(str(file_path), str(target_path))
                except Exception as e:
                    # Fallback: sanitize filename more aggressively and retry
                    safe_name = re.sub(r"[\n\r\t]+", " ", failure_name)
                    safe_name = re.sub(r"\s+", " ", safe_name).strip()
                    safe_name = re.sub(r'[\\/*?:"<>|]', '', safe_name)
                    target_path = failure / safe_name
                    try:
                        shutil.copy2(str(file_path), str(target_path))
                        print(f"      ⚠️ Filename adjusted to: {safe_name}")
                    except Exception as e2:
                        print(f"      ❌ Impossible de copier le fichier vers {target_path}: {e2}")
                # Copier aussi le PDF OCRisé en cas d'échec (pour consultation ultérieure)
                if tmp_pdf and os.path.exists(tmp_pdf):
                    pdf_failure_name = generate_name(inst or "inconnu", obj or "inconnu", failure_date, ".pdf")
                    shutil.copy2(tmp_pdf, str(failure / pdf_failure_name))
                    print(f"  └─ PDF OCRisé copié: {pdf_failure_name}")
                    _temp_files.remove(tmp_pdf) if tmp_pdf in _temp_files else None
                # Exporter le texte vision dans un fichier .txt
                if vision_analysis:
                    vision_text_for_file = vision_analysis.get("text", "") if isinstance(vision_analysis, dict) else vision_analysis
                    vision_structured_for_file = vision_analysis.get("structured", "") if isinstance(vision_analysis, dict) else ""
                    txt_failure_name = generate_name(inst or "inconnu", obj or "inconnu", failure_date, ".txt")
                    txt_path = failure / txt_failure_name
                    with open(txt_path, 'w', encoding='utf-8') as f:
                        if vision_text_for_file:
                            f.write(f"=== TEXTE EXTRAIT PAR VISION ===\n\n{vision_text_for_file}\n")
                        if vision_structured_for_file:
                            f.write(f"\n\n=== ANALYSE STRUCTURÉE VISION ===\n\n{vision_structured_for_file}\n")
                        if text_primary:
                            f.write(f"\n\n=== TEXTE OCR (TESSERACT) ===\n\n{text_primary}\n")
                    print(f"  └─ Texte vision exporté: {txt_failure_name}")
                with open(log_path, 'a', newline='', encoding='utf-8') as f:
                    csv.writer(f).writerow([file_path.name, "Échec", failure_name, inst, obj, date])
                continue
            print("      ✅ Validation OK")
            
            # Renommage
            new_name = generate_name(inst, obj, date, ext)
            new_path = export / new_name
            
            # Pour les PDF, enrichir avec le texte vision si disponible
            if ext == ".pdf" and vision_analysis and PYPDF_AVAILABLE:
                # Enrichir le PDF (original ou searchable) avec les métadonnées vision
                pdf_to_enrich = tmp_pdf if tmp_pdf else str(file_path)
                print("  📝 [ENRICHISSEMENT] Ajout texte vision aux métadonnées PDF...")
                vision_text_for_pdf = vision_analysis.get("text", "") if isinstance(vision_analysis, dict) else vision_analysis
                enriched_pdf = enrich_pdf_with_vision_text(pdf_to_enrich, vision_text_for_pdf, text_primary)
                print("      ✅ Métadonnées enrichies")
                
                pdf_name = generate_name(inst, obj, date, ".pdf")
                shutil.copy2(enriched_pdf, str(export / pdf_name))
                print(f"  📄 PDF final: {pdf_name}")
                
                # Copier aussi le fichier original si c'est une image ou autre format
                if ext != ".pdf":
                    shutil.copy2(str(file_path), str(new_path))
                
                # Nettoyage
                if tmp_pdf and tmp_pdf in _temp_files:
                    _temp_files.remove(tmp_pdf)
                if enriched_pdf != pdf_to_enrich and enriched_pdf in _temp_files:
                    _temp_files.remove(enriched_pdf)
            
            elif tmp_pdf:
                # Image convertie en PDF searchable (sans texte vision)
                pdf_name = generate_name(inst, obj, date, ".pdf")
                shutil.copy2(tmp_pdf, str(export / pdf_name))
                print(f"  📄 PDF searchable: {pdf_name}")
                shutil.copy2(str(file_path), str(new_path))
                _temp_files.remove(tmp_pdf) if tmp_pdf in _temp_files else None
            
            else:
                # Autres formats (DOCX, XLSX, etc.) - copie simple
                shutil.copy2(str(file_path), str(new_path))
            
            # Exporter le texte vision dans un fichier .txt (backup consultable)
            if vision_analysis:
                vision_text_for_file = vision_analysis.get("text", "") if isinstance(vision_analysis, dict) else vision_analysis
                vision_structured_for_file = vision_analysis.get("structured", "") if isinstance(vision_analysis, dict) else ""
                txt_name = generate_name(inst, obj, date, ".txt")
                txt_path = export / txt_name
                with open(txt_path, 'w', encoding='utf-8') as f:
                    if vision_text_for_file:
                        f.write(f"=== TEXTE EXTRAIT PAR VISION ===\n\n{vision_text_for_file}\n")
                    if vision_structured_for_file:
                        f.write(f"\n\n=== ANALYSE STRUCTURÉE VISION ===\n\n{vision_structured_for_file}\n")
                    if text_primary:
                        f.write(f"\n\n=== TEXTE OCR (TESSERACT) ===\n\n{text_primary}\n")
                print(f"  └─ Texte vision exporté: {txt_name}")
            
            print(f"  🎉 EXPORTÉ: {new_name}")
            with open(log_path, 'a', newline='', encoding='utf-8') as f:
                csv.writer(f).writerow([file_path.name, "Succès", new_name, inst, obj, date])
    
    except KeyboardInterrupt:
        print("\n\n⚠️  Interrompu par l'utilisateur")
        cleanup_temp_files()
        sys.exit(130)
    except Exception as e:
        print(f"\n❌ Erreur inattendue: {e}")
        cleanup_temp_files()
        raise
    finally:
        cleanup_temp_files()
        print("\n✅ Exécution terminée.")

if __name__ == "__main__":
    main()
